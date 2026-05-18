"""输出最终扁平子赛道 Excel。"""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 加载数据
data = json.loads(Path("data/final_sub_niches.json").read_text(encoding="utf-8"))

header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "细分赛道"

headers = ["序号", "子赛道名", "原父赛道", "ASIN数", "Top 10 ASIN", "Top 10 关键词"]
for j, h in enumerate(headers, 1):
    cell = ws.cell(1, j, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 加载关键词
import sys
sys.path.insert(0, str(Path.cwd()))
from app.models import SessionLocal
from sqlalchemy import text
db = SessionLocal()

for i, s in enumerate(data["sub_niches"]):
    # 取关键词
    kws = db.execute(text(
        "SELECT keyword FROM sub_niche_kw WHERE sub_niche_id=:sid"
    ), {"sid": s["id"]}).fetchall() if s["id"] else []
    kw_list = [k[0] for k in kws]

    values = [
        i + 1,
        s["name"],
        s["original_parent"],
        s["asin_count"],
        ", ".join(s["asins"][:10]),
        ", ".join(kw_list[:10]) if kw_list else "",
    ]
    for j, v in enumerate(values, 1):
        cell = ws.cell(i + 2, j, v)
        cell.border = thin_border

db.close()

col_widths = [6, 22, 18, 8, 55, 60]
for j, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w

ws.auto_filter.ref = f"A1:F{len(data['sub_niches'])+1}"
ws.freeze_panes = "A2"

out = Path("data/final_sub_niches.xlsx")
wb.save(str(out))
print(f"Saved: {out} — {len(data['sub_niches'])} 个子赛道")
