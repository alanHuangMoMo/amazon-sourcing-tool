"""
细分赛道扩容：ASIN拓词 → 词拓ASIN，双向扩展关键词和产品池。

流程:
  1. 取细分赛道 ASIN，按 ABA 关键词关联数取 TOP5
  2. 凑满 10 个 → 调 keyword_comparison_batch.py 跑 ASIN→P1关键词
  3. 清洗 P1 关键词 → 回写 sub_niche_kw
  4. 汇总所有关键词 → 调 keyword_store_batch.py 跑 关键词→ASIN
  5. 清洗新 ASIN → 回写 final_sub_niche

用法:
  python expand_sub_niche.py --sub-id 17
  python expand_sub_niche.py --sub-id 17 --market CA
"""
import argparse
import json
import re
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

# 项目根
PROJECT_ROOT = Path(__file__).resolve().parent.parent
COMPARISON_SCRIPT = PROJECT_ROOT / "sellersprite_keyword_comparison" / "keyword_comparison_batch.py"
KEYWORD_STORE_SCRIPT = PROJECT_ROOT / "sellersprite_keyword_store" / "keyword_store_batch.py"
UPLOADS_DIR = Path(__file__).resolve().parent / "uploads"

SCRIPT_PYTHON = r"C:\Users\alanh\AppData\Local\Programs\Python\Python313\python.exe"

# 计数器文件
COUNTER_FILE = Path(__file__).resolve().parent / "data" / "sellersprite_export_count.json"


def load_counter():
    if COUNTER_FILE.exists():
        return json.loads(COUNTER_FILE.read_text())
    return {"date": time.strftime("%Y-%m-%d"), "exports": 0}


def save_counter(c):
    COUNTER_FILE.write_text(json.dumps(c))


def check_limit():
    c = load_counter()
    today = time.strftime("%Y-%m-%d")
    if c["date"] != today:
        c = {"date": today, "exports": 0}
        save_counter(c)
    if c["exports"] >= 50:
        print(f"!!! 今日导出已达上限 ({c['exports']}/50) !!!")
        sys.exit(1)
    return c


def increment_counter(c):
    c["exports"] += 1
    save_counter(c)
    print(f"  导出计数: {c['exports']}/50")


def safe_print(text: str):
    """避免 GBK 终端编码报错。"""
    for line in text.splitlines():
        try:
            print(line)
        except UnicodeEncodeError:
            print(line.encode("ascii", errors="replace").decode("ascii")[:200] + "[...]")


def _run_subprocess(cmd: list[str], output_dir: Path, label: str, step: str) -> subprocess.CompletedProcess:
    """统一子进程调用 + 日志。"""
    result = subprocess.run(cmd, capture_output=True, timeout=600,
                            encoding="utf-8", errors="replace")
    # 写日志
    log_file = output_dir / f"{step}_{label}.log"
    log_file.write_text(
        "\n".join(cmd) + "\n\n--- STDOUT ---\n" + result.stdout + "\n--- STDERR ---\n" + result.stderr,
        encoding="utf-8")
    # 打印最后 20 行
    safe_print("\n".join(result.stdout.splitlines()[-20:]))
    if result.returncode != 0:
        safe_print(f"[失败 exit={result.returncode}] 日志: {log_file}")
    return result


def run_comparison(asins: list[str], market: str, output_dir: Path, label: str) -> Path | None:
    """调 keyword_comparison_batch.py"""
    asin_file = output_dir / f"asins_{label}.txt"
    asin_file.write_text("\n".join(asins), encoding="utf-8")

    cmd = [
        SCRIPT_PYTHON, str(COMPARISON_SCRIPT), str(asin_file),
        "--market", market, "--prefix", label, "--output", str(output_dir),
    ]
    print(f"\n=== Step 1: ASIN→关键词 ===")
    print(f"  脚本: {COMPARISON_SCRIPT.name}")
    print(f"  ASIN: {len(asins)} 个")
    result = _run_subprocess(cmd, output_dir, label, "comparison")
    if result.returncode != 0:
        return None

    time.sleep(2)
    p1_files = sorted(output_dir.glob(f"P1-*{label}*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if p1_files:
        return p1_files[0]
    raw_files = sorted(output_dir.glob("CompareKeywords-*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if raw_files:
        print(f"  使用原始导出: {raw_files[0].name}")
        return raw_files[0]
    print("  未找到导出文件")
    safe_print(f"  uploads 目录: {list(output_dir.glob('*.xlsx'))}")
    return None


def parse_p1_excel(xlsx_path: Path) -> dict[str, list[dict]]:
    """解析 P1 Excel → {ASIN: [{keyword, share, search_volume}]}"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    result = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        asin = ws.cell(row, 1).value
        kw = ws.cell(row, 2).value
        share = ws.cell(row, 3).value
        sv = ws.cell(row, 4).value
        if asin and kw:
            result[str(asin).strip()].append({
                "keyword": str(kw).strip(),
                "share": str(share).strip() if share else "",
                "search_volume": str(sv).strip() if sv else "",
            })
    return dict(result)


def run_keyword_store(keywords: list[str], market: str, output_dir: Path, label: str) -> Path | None:
    """调 keyword_store_batch.py"""
    kw_file = output_dir / f"keywords_{label}.txt"
    kw_file.write_text("\n".join(keywords), encoding="utf-8")

    cmd = [
        SCRIPT_PYTHON, str(KEYWORD_STORE_SCRIPT), str(kw_file),
        "--market", market, "--prefix", label, "--output", str(output_dir),
    ]
    print(f"\n=== Step 3: 关键词→ASIN ===")
    print(f"  脚本: {KEYWORD_STORE_SCRIPT.name}")
    result = _run_subprocess(cmd, output_dir, label, "kwstore")
    if result.returncode != 0:
        return None

    time.sleep(2)
    xlsx_files = sorted(output_dir.glob(f"KeywordList-*{label}*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if xlsx_files:
        return xlsx_files[0]
    all_xlsx = sorted(output_dir.glob("KeywordList-*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if all_xlsx:
        print(f"  使用最近导出: {all_xlsx[0].name}")
        return all_xlsx[0]
    safe_print(f"  uploads 目录: {list(output_dir.glob('*.xlsx'))}")
    return None


def parse_kw_store_excel(xlsx_path: Path) -> list[str]:
    """从关键词词库导出中提取前10 ASIN列。"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 找"前10ASIN"列
    header_row = 1
    top10_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val and "10" in str(val) and "ASIN" in str(val):
            top10_col = col
            break

    if top10_col is None:
        print("  未找到前10ASIN列")
        return []

    asins = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, top10_col).value
        if val:
            # 可能是逗号或换行分隔
            for a in re.split(r'[,\n]+', str(val)):
                a = a.strip()
                if len(a) >= 10 and a[0] == 'B':
                    asins.add(a)

    return sorted(asins)


def main():
    parser = argparse.ArgumentParser(description="细分赛道双向扩容")
    parser.add_argument("--sub-id", type=int, required=True, help="子赛道 ID")
    parser.add_argument("--market", default="CA", help="站点")
    args = parser.parse_args()

    from app.models import SessionLocal
    from sqlalchemy import text
    db = SessionLocal()

    # === 0. 加载子赛道数据 ===
    row = db.execute(text(
        "SELECT id, name, asin_count, asins_json FROM final_sub_niche WHERE id=:sid"
    ), {"sid": args.sub_id}).fetchone()
    if not row:
        print(f"子赛道 ID={args.sub_id} 不存在")
        return

    fid, name, asin_cnt, asins_json = row
    asins = json.loads(asins_json)
    print(f"赛道: {name} | {len(asins)} ASIN")
    print(f"ASINs: {', '.join(asins[:5])}{'...' if len(asins) > 5 else ''}")

    # 计算每个 ASIN 的 ABA 关键词关联数
    asin_kw_count = {}
    for a in asins:
        cnt = db.execute(text("""
            SELECT COUNT(*) FROM aba_report WHERE domain='CA'
            AND (asin_1=:a OR asin_2=:a OR asin_3=:a)
        """), {"a": a}).scalar()
        asin_kw_count[a] = cnt or 0

    sorted_asins = sorted(asins, key=lambda a: -asin_kw_count[a])
    print(f"\nABA 关键词关联数排名:")
    for a in sorted_asins:
        print(f"  {a}: {asin_kw_count[a]} kw")

    # 取 TOP5，凑满 10
    top5 = sorted_asins[:5]
    remaining = [a for a in sorted_asins if a not in top5]
    query_asins = top5 + remaining[:10 - len(top5)]
    print(f"\n查询 ASIN ({len(query_asins)} 个): {', '.join(query_asins)}")

    # 检查导出次数
    counter = check_limit()

    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    now_str = time.strftime("%Y%m%d-%H%M%S")
    label = f"{name}-{now_str}"

    # === Step 1: ASIN→关键词 ===
    p1_file = run_comparison(query_asins, args.market, UPLOADS_DIR, label)
    if not p1_file:
        print("Step 1 失败，退出")
        return
    increment_counter(counter)

    # === Step 2: 清洗 + 回写关键词 ===
    print(f"\n=== Step 2: 清洗关键词 ===")
    p1_data = parse_p1_excel(p1_file)
    total_p1_kw = 0
    new_kw_count = 0

    for asin, entries in p1_data.items():
        print(f"  {asin}: {len(entries)} P1 关键词")
        total_p1_kw += len(entries)

    # 收集所有 P1 关键词，去重写入 sub_niche_kw
    all_new_kws = set()
    for asin, entries in p1_data.items():
        for e in entries:
            all_new_kws.add(e["keyword"])

    # 查已有关键词
    existing_kws = set()
    for r in db.execute(text(
        "SELECT keyword FROM sub_niche_kw WHERE sub_niche_id=:sid"
    ), {"sid": args.sub_id}).fetchall():
        existing_kws.add(r[0])

    truly_new = all_new_kws - existing_kws
    for kw in truly_new:
        db.execute(text(
            "INSERT INTO sub_niche_kw (sub_niche_id, keyword) VALUES (:sid, :kw)"
        ), {"sid": args.sub_id, "kw": kw})

    db.commit()
    print(f"\nP1 关键词: {len(all_new_kws)} 去重")
    print(f"已有: {len(existing_kws)}, 新增: {len(truly_new)}")
    print(f"回写完成，sub_niche_kw 新增 {len(truly_new)} 条")

    # === Step 3: 关键词→ASIN ===
    # 汇总所有关键词
    all_kws = list(existing_kws | all_new_kws)
    print(f"\n全量关键词: {len(all_kws)} 个")

    if len(all_kws) > 2000:
        print(f"  超过 2000，截取前 2000")
        all_kws = all_kws[:2000]

    counter = load_counter()
    counter = check_limit()  # 重新检查
    kw_file = run_keyword_store(all_kws, args.market, UPLOADS_DIR, label)
    if not kw_file:
        print("Step 3 失败，退出")
        return
    increment_counter(counter)

    # === Step 4: 清洗 + 回写 ASIN ===
    print(f"\n=== Step 4: 清洗 ASIN ===")
    new_asins = parse_kw_store_excel(kw_file)
    print(f"从关键词词库提取: {len(new_asins)} 个新 ASIN")

    # 去重已有
    existing_asins = set(asins)
    truly_new_asins = [a for a in new_asins if a not in existing_asins]
    print(f"已有: {len(existing_asins)}, 新增: {len(truly_new_asins)}")

    # 更新 final_sub_niche
    all_asins = sorted(existing_asins | set(truly_new_asins))
    db.execute(text(
        "UPDATE final_sub_niche SET asin_count=:c, asins_json=:aj WHERE id=:sid"
    ), {"c": len(all_asins), "aj": json.dumps(all_asins, ensure_ascii=False), "sid": args.sub_id})

    # 更新 sub_niche 也保持一致
    db.execute(text(
        "UPDATE sub_niche SET size=:c WHERE id=:sid"
    ), {"c": len(all_asins), "sid": args.sub_id})

    for a in truly_new_asins:
        try:
            db.execute(text(
                "INSERT INTO final_sub_niche_asin (sub_niche_id, asin) VALUES (:sid, :a)"
            ), {"sid": args.sub_id, "a": a})
        except Exception:
            pass
        try:
            db.execute(text(
                "INSERT INTO sub_niche_asin (sub_niche_id, asin) VALUES (:sid, :a)"
            ), {"sid": args.sub_id, "a": a})
        except Exception:
            pass

    db.commit()

    print(f"\n=== 完成 ===")
    print(f"赛道: {name}")
    print(f"ASIN: {len(asins)} → {len(all_asins)} (+{len(truly_new_asins)})")
    print(f"关键词: {len(existing_kws)} → {len(existing_kws) + len(truly_new)} (+{len(truly_new)})")
    print(f"今日导出: {counter['exports']}/50")

    # 保存中间文件路径供查看
    print(f"\n文件:")
    print(f"  P1: {p1_file}")
    print(f"  KeywordList: {kw_file}")

    db.close()


if __name__ == "__main__":
    main()
