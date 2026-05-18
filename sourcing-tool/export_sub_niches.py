"""导出细分赛道全景 Excel 报告。"""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from app.models import init_db, SessionLocal
from sqlalchemy import text

db = SessionLocal()

# === 1. 加载子赛道数据 ===
rows = db.execute(text("""
    SELECT sn.id, sn.parent_tag, sn.name, sn.belongs_to_niche, sn.size, sn.keyword_count,
           sn.merged_niche_id
    FROM sub_niche sn
    ORDER BY sn.parent_tag, sn.size DESC
""")).fetchall()

# 加载每个子赛道的 ASIN 和关键词
sub_asins = {}
sub_kws = {}
asin_rows = db.execute(text("SELECT sub_niche_id, asin FROM sub_niche_asin")).fetchall()
for sid, asin in asin_rows:
    sub_asins.setdefault(sid, []).append(asin)
kw_rows = db.execute(text("SELECT sub_niche_id, keyword FROM sub_niche_kw")).fetchall()
for sid, kw in kw_rows:
    sub_kws.setdefault(sid, []).append(kw)

# === 2. 拉卖家精灵产品数据 ===
all_asins = set()
for asins in sub_asins.values():
    all_asins.update(asins)

placeholders = ",".join(f":a{i}" for i in range(len(all_asins)))
params = {f"a{i}": a for i, a in enumerate(all_asins)}
prod_rows = db.execute(text(f"""
    SELECT asin, price, ratings, ratings_count, monthly_sales, monthly_revenue,
           profit_rate, fba_fee, seller_count, online_date, brand, title
    FROM sellersprite_product WHERE domain='CA' AND asin IN ({placeholders})
"""), params).fetchall()

prod_map = {}
for r in prod_rows:
    prod_map[r[0]] = {
        "price": r[1] or 0,
        "ratings": r[2] or 0,
        "ratings_count": r[3] or 0,
        "monthly_sales": r[4] or 0,
        "monthly_revenue": r[5] or 0,
        "profit_rate": r[6] or 0,
        "fba_fee": r[7] or 0,
        "seller_count": r[8] or 0,
        "online_date": r[9] or "",
        "brand": r[10] or "",
        "title": r[11] or "",
    }

db.close()

# === 3. 计算每个子赛道聚合指标 ===
def calc_metrics(asins):
    prices = []
    ratings = []
    review_counts = []
    sales = []
    revenues = []
    profit_rates = []
    fba_fees = []
    seller_counts = []
    brands = set()

    for a in asins:
        p = prod_map.get(a)
        if not p:
            continue
        if p["price"] > 0:
            prices.append(p["price"])
        if p["ratings"] > 0:
            ratings.append(p["ratings"])
        if p["ratings_count"] > 0:
            review_counts.append(p["ratings_count"])
        if p["monthly_sales"] > 0:
            sales.append(p["monthly_sales"])
        if p["monthly_revenue"] > 0:
            revenues.append(p["monthly_revenue"])
        if p["profit_rate"] > 0:
            profit_rates.append(p["profit_rate"])
        if p["fba_fee"] > 0:
            fba_fees.append(p["fba_fee"])
        if p["seller_count"] > 0:
            seller_counts.append(p["seller_count"])
        if p["brand"]:
            brands.add(p["brand"])

    return {
        "data_asins": len([a for a in asins if a in prod_map]),
        "avg_price": round(sum(prices)/len(prices), 2) if prices else 0,
        "median_price": round(sorted(prices)[len(prices)//2], 2) if prices else 0,
        "avg_ratings": round(sum(ratings)/len(ratings), 2) if ratings else 0,
        "avg_reviews": int(sum(review_counts)/len(review_counts)) if review_counts else 0,
        "total_monthly_sales": int(sum(sales)),
        "total_monthly_revenue": int(sum(revenues)),
        "avg_profit_rate": round(sum(profit_rates)/len(profit_rates)*100, 1) if profit_rates else 0,
        "avg_fba_fee": round(sum(fba_fees)/len(fba_fees), 2) if fba_fees else 0,
        "avg_seller_count": round(sum(seller_counts)/len(seller_counts), 1) if seller_counts else 0,
        "brand_count": len(brands),
    }

# === 4. 写 Excel ===
wb = openpyxl.Workbook()

# --- Sheet 1: 全景总览 ---
ws1 = wb.active
ws1.title = "细分赛道全景"

header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
noise_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

headers = [
    "父赛道", "子赛道名", "状态", "ASIN数", "关键词数",
    "有数据ASIN", "均价(CAD)", "中位价(CAD)", "均评分",
    "均评论数", "月总销量", "月总营收(CAD)", "均利润率%",
    "均FBA费", "均卖家数", "品牌数",
    "Top 5 ASIN", "Top 10 关键词",
]

for j, h in enumerate(headers, 1):
    cell = ws1.cell(1, j, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

row_num = 2
for r in rows:
    sid, tag, name, belongs, size, kw_cnt, mnid = r
    asins = sub_asins.get(sid, [])
    kws = sub_kws.get(sid, [])
    metrics = calc_metrics(asins)

    status = "有效" if belongs else "⚠ 噪音"
    top5_asin = ", ".join(asins[:5])
    top10_kw = ", ".join(kws[:10])

    values = [
        tag, name or "(未命名)", status, size, kw_cnt,
        metrics["data_asins"], metrics["avg_price"], metrics["median_price"],
        metrics["avg_ratings"], metrics["avg_reviews"],
        metrics["total_monthly_sales"], metrics["total_monthly_revenue"],
        metrics["avg_profit_rate"], metrics["avg_fba_fee"],
        metrics["avg_seller_count"], metrics["brand_count"],
        top5_asin, top10_kw,
    ]

    for j, v in enumerate(values, 1):
        cell = ws1.cell(row_num, j, v)
        cell.border = thin_border
        if not belongs:
            cell.fill = noise_fill
        if isinstance(v, float):
            cell.number_format = "0.00"

    row_num += 1

# 列宽
col_widths = [18, 16, 8, 8, 8, 10, 10, 10, 8, 10, 10, 14, 10, 10, 10, 8, 50, 60]
for j, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(j)].width = w

ws1.auto_filter.ref = f"A1:R{row_num-1}"
ws1.freeze_panes = "A2"

# --- Sheet 2: 按父赛道汇总 ---
ws2 = wb.create_sheet("父赛道汇总")

# 按父赛道聚合
tag_stats = defaultdict(lambda: {"sub_count": 0, "total_asin": 0, "total_kw": 0,
                                   "valid_subs": 0, "noise_subs": 0, "names": [],
                                   "all_asins": set()})
for r in rows:
    sid, tag, name, belongs, size, kw_cnt, mnid = r
    ts = tag_stats[tag]
    ts["sub_count"] += 1
    ts["total_asin"] += size
    ts["total_kw"] += kw_cnt
    if belongs:
        ts["valid_subs"] += 1
        ts["names"].append(name or tag)
    else:
        ts["noise_subs"] += 1
    ts["all_asins"].update(sub_asins.get(sid, []))

summary_headers = ["父赛道", "子赛道数", "有效", "噪音", "总ASIN", "总KW",
                   "唯一ASIN", "有数据ASIN", "均价", "均评论", "月总销量", "子赛道名"]
for j, h in enumerate(summary_headers, 1):
    cell = ws2.cell(1, j, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

row_num = 2
for tag, ts in sorted(tag_stats.items(), key=lambda x: -x[1]["total_asin"]):
    metrics = calc_metrics(list(ts["all_asins"]))
    values = [
        tag, ts["sub_count"], ts["valid_subs"], ts["noise_subs"],
        ts["total_asin"], ts["total_kw"],
        len(ts["all_asins"]), metrics["data_asins"],
        metrics["avg_price"], metrics["avg_reviews"],
        metrics["total_monthly_sales"],
        " | ".join(ts["names"][:5]),
    ]
    for j, v in enumerate(values, 1):
        cell = ws2.cell(row_num, j, v)
        cell.border = thin_border
    row_num += 1

for j, w in enumerate([18, 10, 8, 8, 10, 10, 10, 10, 10, 10, 12, 60], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.auto_filter.ref = f"A1:L{row_num-1}"
ws2.freeze_panes = "A2"

# --- Sheet 3: 产品数据明细 ---
ws3 = wb.create_sheet("产品数据")

prod_headers = ["父赛道", "子赛道", "ASIN", "标题", "品牌", "价格(CAD)",
                "评分", "评论数", "月销量", "月营收", "利润率", "FBA费",
                "卖家数", "上架日期"]
for j, h in enumerate(prod_headers, 1):
    cell = ws3.cell(1, j, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

row_num = 2
for r in rows:
    sid, tag, name, belongs, size, kw_cnt, mnid = r
    if not belongs:
        continue
    for asin in sub_asins.get(sid, []):
        p = prod_map.get(asin, {})
        values = [
            tag, name or tag, asin,
            p.get("title", "")[:120],
            p.get("brand", ""),
            p.get("price", ""),
            p.get("ratings", ""),
            p.get("ratings_count", ""),
            p.get("monthly_sales", ""),
            p.get("monthly_revenue", ""),
            p.get("profit_rate", ""),
            p.get("fba_fee", ""),
            p.get("seller_count", ""),
            p.get("online_date", ""),
        ]
        for j, v in enumerate(values, 1):
            ws3.cell(row_num, j, v).border = thin_border
        row_num += 1

for j, w in enumerate([18, 16, 12, 60, 15, 10, 6, 8, 8, 10, 8, 8, 8, 12], 1):
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.auto_filter.ref = f"A1:N{row_num-1}"
ws3.freeze_panes = "C2"

# 保存
out_path = Path("data/sub_niches_report.xlsx")
wb.save(str(out_path))
print(f"Saved: {out_path}")
print(f"  Sheet 1: 细分赛道全景 — {sum(1 for r in rows)} 个子赛道")
print(f"  Sheet 2: 父赛道汇总 — {len(tag_stats)} 个父赛道")
print(f"  Sheet 3: 产品数据明细")
