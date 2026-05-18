"""
批量 P1 自然排名关键词扩容：为所有子赛道头部 ASIN 跑流量词对比，
提取第1页自然排名关键词，回填到 final_sub_niche_kw。

流程:
  1. 每子赛道取前 3 个 ASIN
  2. 每 10 个 ASIN 一批 → 调 keyword_comparison_batch.py
  3. 解析自然排名 Excel → 提取第1页关键词
  4. 通过 ASIN 映射回 sub_niche_id → INSERT OR IGNORE

用法:
  python batch_p1_expand.py              # 从头跑
  python batch_p1_expand.py --resume     # 续跑（跳过已完成批次）
  python batch_p1_expand.py --dry-run    # 只看计划不执行
"""

import argparse
import json
import os
import re
import sqlite3
import subprocess
import sys
import time
import traceback
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
COMPARISON_SCRIPT = PROJECT_ROOT / "sellersprite_keyword_comparison" / "keyword_comparison_batch.py"
SCRIPT_PYTHON = r"C:\Users\alanh\AppData\Local\Programs\Python\Python313\python.exe"
DATA_DIR = Path(__file__).resolve().parent / "data"
BATCH_DIR = DATA_DIR / "p1_batches"
OUTPUT_DIR = DATA_DIR / "p1_outputs"

BATCH_SIZE = 10
TOP_ASINS_PER_NICHE = 3
MARKET = "CA"
DIMENSION = "natural_rank"
PROFILE_DIR = Path.home() / ".sellersprite-comparison-profile"


EXPORT_QUOTA = 50
EXPORT_COUNT_FILE = DATA_DIR / "sellersprite_export_count.json"


def _cleanup_chrome():
    """杀掉所有 Chrome 进程"""
    import subprocess as sp
    try:
        sp.run(["taskkill", "/F", "/IM", "chrome.exe"],
               capture_output=True, timeout=30)
    except Exception:
        pass


EXPORT_KEY = "comparison"  # 流量词对比


def load_export_count() -> dict:
    """读取今日导出次数，跨日自动归零"""
    today = time.strftime("%Y-%m-%d")
    if EXPORT_COUNT_FILE.exists():
        data = json.loads(EXPORT_COUNT_FILE.read_text())
        if data.get("date") != today:
            return {"date": today, "comparison": 0, "keyword_store": 0}
        data.setdefault(EXPORT_KEY, 0)
        return data
    return {"date": today, "comparison": 0, "keyword_store": 0}


def increment_export_count():
    """导出计数 +1"""
    data = load_export_count()
    data[EXPORT_KEY] = data.get(EXPORT_KEY, 0) + 1
    EXPORT_COUNT_FILE.parent.mkdir(parents=True, exist_ok=True)
    EXPORT_COUNT_FILE.write_text(json.dumps(data))


def check_quota() -> bool:
    """返回 True 表示额度充足，False 表示今日已用完"""
    data = load_export_count()
    cnt = data.get(EXPORT_KEY, 0)
    remain = EXPORT_QUOTA - cnt
    print(f"今日 {EXPORT_KEY}: {cnt}/{EXPORT_QUOTA} (剩余 {remain})")
    if remain <= 0:
        print(f"!!! 今日额度用完，请明天继续")
        return False
    return True


def load_db():
    db_path = DATA_DIR / "sourcing.db"
    db = sqlite3.connect(str(db_path))
    db.execute("PRAGMA journal_mode=WAL")
    return db


def get_batch_asins(db) -> list[list[tuple[str, int]]]:
    """每个子赛道取前 3 个 ASIN，返回 [(asin, sub_niche_id), ...] 列表，按每批 10 个分好组"""
    rows = db.execute("""
        SELECT fsa.sub_niche_id, fsa.asin
        FROM final_sub_niche_asin fsa
        JOIN final_sub_niche fs ON fs.id = fsa.sub_niche_id
        ORDER BY fsa.sub_niche_id, fsa.id
    """).fetchall()

    # 按 sub_niche_id 分组
    niche_asins = defaultdict(list)
    for sid, asin in rows:
        niche_asins[sid].append(asin)

    # 每子赛道取前 N
    selected = []
    for sid in sorted(niche_asins):
        asins = niche_asins[sid]
        for a in asins[:TOP_ASINS_PER_NICHE]:
            selected.append((a, sid))

    # 分批
    batches = []
    for i in range(0, len(selected), BATCH_SIZE):
        batches.append(selected[i:i + BATCH_SIZE])

    return batches


def run_comparison_batch(input_asins: list[tuple[str, int]], batch_idx: int) -> tuple[Path | None, list[str]]:
    """调 keyword_comparison_batch.py，返回 (P1报告路径, 输入ASIN顺序列表用于变体位置匹配)"""
    asins = [a for a, _ in input_asins]
    batch_file = BATCH_DIR / f"batch_{batch_idx:04d}.txt"
    batch_file.parent.mkdir(parents=True, exist_ok=True)
    batch_file.write_text("\n".join(asins), encoding="utf-8")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    cmd = [
        SCRIPT_PYTHON, str(COMPARISON_SCRIPT),
        str(batch_file),
        "--market", MARKET,
        "--dimension", DIMENSION,
        "--output", str(OUTPUT_DIR),
        "--prefix", f"BATCH-{batch_idx:04d}",
        "--top", "10000",  # 打印全部 P1 关键词
    ]

    print(f"\n{'='*60}")
    print(f"批次 {batch_idx}: {len(asins)} ASINs")
    for a in asins:
        print(f"  {a}")
    print(f"{'='*60}")

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=600,
            encoding="utf-8",
            errors="replace",
            cwd=str(PROJECT_ROOT / "sellersprite_keyword_comparison"),
            env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        stdout = result.stdout
        stderr = result.stderr
        if stderr:
            print(f"  [stderr]: {stderr[:500]}")

        # 只打印尾部（关键信息）
        lines = stdout.splitlines()
        for line in lines[-30:]:
            print(line)

        if result.returncode != 0:
            print(f"  批次 {batch_idx} 失败 (exit {result.returncode})")
            return None, asins

        # 优先找 P1 报告（已清洗好），再 fallback 原始文件
        p1_report = None
        raw_file = None
        for line in stdout.splitlines():
            if "报告:" in line:
                path_str = line.split(":", 1)[-1].strip()
                p = Path(path_str)
                if p.exists():
                    p1_report = p
            if "原始:" in line:
                path_str = line.split(":", 1)[-1].strip()
                p = Path(path_str)
                if p.exists():
                    raw_file = p

        target = p1_report or raw_file
        if target:
            return target, asins

        # fallback
        xlsx_files = sorted(OUTPUT_DIR.glob("P1-Organic-*.xlsx"),
                           key=lambda f: f.stat().st_mtime, reverse=True)
        if xlsx_files:
            return xlsx_files[0], asins

        print(f"  批次 {batch_idx}: 找不到输出文件")
        return None, asins

    except subprocess.TimeoutExpired:
        print(f"  批次 {batch_idx} 超时")
        return None, asins


def parse_p1_report(xlsx_path: Path) -> dict[str, list[str]]:
    """解析 P1 自然排名报告 (ASIN|关键词|自然排名|页码|月搜索量)，返回 {asin: [keyword, ...]}"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 读表头确认列
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v:
            headers[col] = str(v).strip()

    asin_col = kw_col = None
    for c, h in headers.items():
        if h == "ASIN":
            asin_col = c
        elif h == "关键词":
            kw_col = c

    if not asin_col or not kw_col:
        # 兼容旧格式：尝试解析原始导出
        print("  P1报告格式不匹配，尝试解析原始导出...")
        return _parse_raw_export(xlsx_path)

    result = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        asin = ws.cell(row, asin_col).value
        kw = ws.cell(row, kw_col).value
        if asin and kw:
            result[str(asin).strip()].append(str(kw).strip())

    return dict(result)


def _parse_raw_export(xlsx_path: Path) -> dict[str, list[str]]:
    """兼容：解析原始自然排名导出"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    asin_cols = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v and len(str(v)) >= 10 and str(v)[:2] == "B0":
            asin_cols.append((str(v).strip(), col, col + 1))

    if not asin_cols:
        return {}

    result = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        kw = ws.cell(row, 1).value
        if not kw:
            continue
        kw = str(kw).strip()
        for asin, rank_col, page_col in asin_cols:
            page_val = ws.cell(row, page_col).value
            if page_val and "第1页" in str(page_val):
                result[asin].append(kw)

    return dict(result)


def store_p1_keywords(db, asin_kw_map: dict[str, list[str]],
                      asin_niche_map: dict[str, set[int]],
                      input_asins: list[str]) -> tuple[int, int]:
    """回填 P1 关键词到 final_sub_niche_kw。
    变体 ASIN 通过「缺失输入 ASIN」匹配：输出中不在 DB 的 ASIN 视为变体，
    按顺序对应到输入中没出现在输出里的 ASIN。
    """
    inserted = 0
    skipped = 0
    output_asins = list(asin_kw_map.keys())
    output_set = set(output_asins)

    # 找出变体映射：输出中不在 DB 的 vs 输入中没出现在输出的
    unmatched_outputs = [a for a in output_asins if a not in asin_niche_map]
    unmatched_inputs = [a for a in input_asins if a not in output_set]

    variant_map = {}  # variant_asin → niche_ids
    for i, v in enumerate(unmatched_outputs):
        if i < len(unmatched_inputs):
            orig = unmatched_inputs[i]
            variant_map[v] = asin_niche_map.get(orig, set())

    for asin, keywords in asin_kw_map.items():
        niche_ids = asin_niche_map.get(asin) or variant_map.get(asin)
        if not niche_ids:
            continue
        for kw in keywords:
            for sid in niche_ids:
                try:
                    db.execute(
                        "INSERT INTO final_sub_niche_kw (sub_niche_id, keyword) VALUES (?, ?)",
                        (sid, kw),
                    )
                    inserted += 1
                except sqlite3.IntegrityError:
                    skipped += 1
    db.commit()
    return inserted, skipped


def save_progress(batch_idx: int, total: int):
    """记录完成进度。只允许递增，防止并发 session 互相覆盖"""
    progress_file = DATA_DIR / "p1_expand_progress.json"
    data = load_progress()
    if data and data.get("last_completed_batch", -1) >= batch_idx:
        return
    progress_file.write_text(json.dumps({
        "last_completed_batch": batch_idx,
        "total_batches": total,
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }))


def load_progress():
    progress_file = DATA_DIR / "p1_expand_progress.json"
    if progress_file.exists():
        return json.loads(progress_file.read_text())
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--max-batches", type=int, default=0, help="达到指定批数后自动停止（0=不限）")
    parser.add_argument("--target-batches", type=int, default=50, help="目标批数，用于最终统计提示")
    parser.add_argument("--batches", type=int, default=0, help="本次只跑 N 批，完成后自动停（0=不限）")
    args = parser.parse_args()

    db = load_db()

    # 1. 准备批次
    batches = get_batch_asins(db)
    total_asins = sum(len(b) for b in batches)
    print(f"共 {len(batches)} 批次, {total_asins} ASINs")

    # 2. ASIN → sub_niche_id 映射
    asin_niche_map = defaultdict(set)
    for batch in batches:
        for asin, sid in batch:
            asin_niche_map[asin].add(sid)

    if args.dry_run:
        print("\n=== 批次预览 ===")
        for i, batch in enumerate(batches[:5]):
            print(f"\n批次 {i}:")
            for asin, sid in batch:
                name = db.execute("SELECT name FROM final_sub_niche WHERE id=?", (sid,)).fetchone()
                n = name[0] if name else "?"
                print(f"  {asin} ← {n} (ID:{sid})")
        print(f"\n... 共 {len(batches)} 批次")
        return

    # 3. 确定起始批次
    start_batch = 0
    if args.resume:
        prog = load_progress()
        if prog:
            start_batch = prog["last_completed_batch"] + 1
            print(f"续跑: 从批次 {start_batch} 开始")

    # 4. 跑批次
    total_inserted = 0
    total_skipped = 0
    failed_batches = []
    quota_exhausted = False

    for i in range(start_batch, len(batches)):
        batch = batches[i]

        # 本次只跑 N 批，交棒给下个 session
        if args.batches and (i - start_batch) >= args.batches:
            print(f"本次已跑 {args.batches} 批，交棒给下个 session")
            break

        # 达到目标批数自动停
        if args.max_batches and i >= args.max_batches:
            print(f"已达到目标 {args.max_batches} 批，停止")
            save_progress(i - 1, len(batches))
            break

        # 检查额度
        if not check_quota():
            quota_exhausted = True
            break

        try:
            xlsx_path, input_order = run_comparison_batch(batch, i)
            # 导出成功就计数（不管后续入库是否成功）
            increment_export_count()

            if not xlsx_path:
                print(f"!!! 批次 {i} 失败，跳过继续")
                failed_batches.append(i)
                save_progress(i, len(batches))
                _cleanup_chrome()
                continue

            # 解析 P1 报告
            asin_kw_map = parse_p1_report(xlsx_path)
            total_kws = sum(len(v) for v in asin_kw_map.values())
            print(f"  解析: {len(asin_kw_map)} ASINs, {total_kws} P1 关键词")

            # 入库
            inserted, skipped = store_p1_keywords(db, asin_kw_map, asin_niche_map, input_order)
            total_inserted += inserted
            total_skipped += skipped
            print(f"  入库: +{inserted} 新词, {skipped} 已存在")

            save_progress(i, len(batches))
            _cleanup_chrome()
        except Exception:
            tb = traceback.format_exc()
            print(f"!!! 批次 {i} 异常: {tb[:500]}")
            err_log = DATA_DIR / "p1_expand_errors.log"
            with open(err_log, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*60}\n")
                f.write(f"批次 {i} 崩溃 @ {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"{'='*60}\n")
                f.write(tb)
                f.write("\n")
            failed_batches.append(i)
            save_progress(i, len(batches))
            _cleanup_chrome()
            continue

        # 每 10 批休息一下
        if (i + 1) % 10 == 0:
            print(f"\n--- 已完成 {i+1}/{len(batches)} 批次, 新增 {total_inserted} 词 ---")

    # 5. 最终统计
    print(f"\n{'='*60}")
    if quota_exhausted:
        data = load_export_count()
        print(f"今日额度用完 ({data[EXPORT_KEY]}/{EXPORT_QUOTA})，已暂停")
        print(f"当前进度: {load_progress()['last_completed_batch']+1}/{len(batches)} 批")
        print(f"明天 cron 自动续跑，或手动: python batch_p1_expand.py --resume")
    else:
        print(f"完成! 新增关键词: {total_inserted}, 跳过(已存在): {total_skipped}")
    if failed_batches:
        print(f"失败批次 ({len(failed_batches)}): {failed_batches}")
    new_total = db.execute("SELECT COUNT(*) FROM final_sub_niche_kw").fetchone()[0]
    new_unique = db.execute("SELECT COUNT(DISTINCT keyword) FROM final_sub_niche_kw").fetchone()[0]
    print(f"final_sub_niche_kw 总计: {new_total} 行, {new_unique} 唯一关键词")
    db.close()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        tb = traceback.format_exc()
        print(f"FATAL: {tb[:1000]}")
        err_log = DATA_DIR / "p1_expand_errors.log"
        with open(err_log, "a", encoding="utf-8") as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"FATAL 崩溃 @ {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*60}\n")
            f.write(tb)
            f.write("\n")
        sys.exit(1)
