"""FastAPI 路由 — 上传、流程、结果、导出。"""
import json
import os
import re
import asyncio as _asyncio
import uuid as _uuid
import threading
from pathlib import Path
from io import BytesIO

from fastapi import APIRouter, UploadFile, File, Form, Request, Query
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

from .models import SessionLocal, Candidate, Batch, AsinKeyword, init_db, AbaReport, SellerspriteKeyword, SellerspriteProduct, WordRoot, NicheTrack, Project
from .pipeline import run_pipeline, PipelineProgress
from .aba_processor import parse_aba_excel, detect_aba_format, parse_metadata
from .auth import check_auth, set_auth_cookie

router = APIRouter()

BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# 启动时加载已保存的探索器配置
try:
    from .explorer_config import load as _load_explorer_cfg, apply_to_module as _apply_cfg
    _apply_cfg(_load_explorer_cfg())
except Exception:
    pass  # 配置文件不存在或损坏时使用默认值

# 内存中存进度状态（简单实现，生产可换 Redis）
_pipeline_states: dict[str, dict] = {}


def _pipeline_thread(batch_id: str, filepath: str, domain: int, config: dict, mock: bool, data_source: str = "sellersprite"):
    """后台线程运行 pipeline。"""
    progress = PipelineProgress()
    _pipeline_states[batch_id] = {"status": "running", "progress": progress}
    try:
        result = run_pipeline(
            filepath=filepath,
            domain=domain,
            config=config,
            mock=mock,
            progress=progress,
            data_source=data_source,
        )
        _pipeline_states[batch_id] = {"status": "done", "result": result, "progress": progress}
    except Exception as e:
        _pipeline_states[batch_id] = {"status": "failed", "error": str(e), "progress": progress}


# ── 认证路由 ────────────────────────────────────

@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """登录页（不依赖 base.html）。"""
    if check_auth(request):
        return RedirectResponse(url="/", status_code=302)
    return templates.TemplateResponse(request, "login.html")


@router.post("/api/auth/login")
async def api_auth_login(request: Request):
    """验证密码，设 cookie。"""
    from .auth import APP_PASSWORD, _hash
    body = await request.json()
    password = body.get("password", "")
    if password == APP_PASSWORD:
        response = RedirectResponse(url="/", status_code=302)
        set_auth_cookie(response)
        return response
    return HTMLResponse(status_code=401)


@router.get("/api/auth/logout")
async def api_auth_logout():
    """清除认证 cookie。"""
    from .auth import COOKIE_NAME
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie(COOKIE_NAME)
    return response


# ── 页面路由 ────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def index_page(request: Request):
    """项目列表首页。"""
    return templates.TemplateResponse(request, "project-list.html")


@router.get("/pipeline", response_class=HTMLResponse)
async def pipeline_page(request: Request):
    """旧版 pipeline 页（保留兼容）。"""
    return templates.TemplateResponse(request, "index.html")


@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard_redirect(request: Request):
    """旧版 /dashboard → 重定向到首页。"""
    return RedirectResponse(url="/", status_code=302)


@router.get("/projects/new", response_class=HTMLResponse)
async def project_create_page(request: Request):
    """新建项目页（上传流程）。"""
    return templates.TemplateResponse(request, "project-create.html")


@router.get("/projects/{project_id}", response_class=HTMLResponse)
async def project_kanban_page(request: Request, project_id: int):
    """项目看板页。"""
    return templates.TemplateResponse(request, "project-kanban.html", {"project_id": project_id})


@router.post("/upload", response_class=HTMLResponse)
async def handle_upload(request: Request, file: UploadFile = File(...)):
    """上传 ABA 报表，返回预览和配置页。"""
    # 保存文件
    ext = file.filename.rsplit(".", 1)[-1].lower()
    filepath = UPLOAD_DIR / f"aba_upload.{ext}"
    content = await file.read()
    filepath.write_bytes(content)

    # 解析预览
    try:
        df = parse_aba_excel(str(filepath))
        col_map = detect_aba_format(df)
        preview = {
            "rows": len(df),
            "keyword_col": col_map.get("keyword", "?"),
            "asin_cols": col_map.get("asin_cols", []),
            "click_cols": col_map.get("click_cols", []),
            "conv_cols": col_map.get("conv_cols", []),
            "brand_cols": col_map.get("brand_cols", []),
        }
    except Exception as e:
        preview = {"error": str(e)}

    return templates.TemplateResponse(request, "pipeline.html", {
        "filepath": str(filepath),
        "filename": file.filename,
        "preview": preview,
    })


@router.post("/run", response_class=HTMLResponse)
async def start_pipeline(
    request: Request,
    filepath: str = Form(...),
    domain: int = Form(1),
    conv_index_min: float = Form(1.0),
    share_max: float = Form(50.0),
    mock: bool = Form(True),
):
    """启动 pipeline 后台线程，返回进度页。"""
    batch_id = f"batch_{os.urandom(4).hex()}"
    config = {
        "conv_index_min": conv_index_min,
        "share_max": share_max,
    }

    thread = threading.Thread(
        target=_pipeline_thread,
        args=(batch_id, filepath, domain, config, mock),
        daemon=True,
    )
    thread.start()

    return templates.TemplateResponse(request, "progress.html", {
        "batch_id": batch_id,
    })


@router.get("/progress/{batch_id}", response_class=HTMLResponse)
async def poll_progress(request: Request, batch_id: str):
    """htmx polling 端点，返回进度 HTML 片段。"""
    state = _pipeline_states.get(batch_id, {})
    if not state:
        return HTMLResponse("<div class='text-red-500'>批次未找到</div>")

    if state["status"] == "running":
        p = state["progress"]
        return templates.TemplateResponse(request, "_progress_bar.html", {
            "percent": p.percent,
            "message": p.message,
            "step": p.current_step,
            "logs": p.logs[-10:],
            "done": False,
        })

    elif state["status"] == "done":
        result = state["result"]
        return templates.TemplateResponse(request, "_progress_bar.html", {
            "percent": 100,
            "message": f"完成! {result['candidate_count']} 个候选 ASIN",
            "step": "完成",
            "logs": state["progress"].logs[-10:] if state.get("progress") else [],
            "done": True,
            "batch_id": batch_id,
        })

    else:
        return templates.TemplateResponse(request, "_progress_bar.html", {
            "percent": 0,
            "message": f"失败: {state.get('error', '未知错误')}",
            "step": "失败",
            "logs": state["progress"].logs[-10:] if state.get("progress") else [],
            "done": True,
            "error": state.get("error"),
        })


@router.get("/results/{batch_id}", response_class=HTMLResponse)
async def results_page(request: Request, batch_id: str):
    """候选清单结果页。"""
    state = _pipeline_states.get(batch_id, {})
    result = state.get("result", {})

    return templates.TemplateResponse(request, "results.html", {
        "batch_id": batch_id,
        "candidates": result.get("candidates", []),
        "summary": result.get("summary", {}),
        "candidate_count": result.get("candidate_count", 0),
    })


@router.get("/api/candidates/{batch_id}")
async def api_candidates(batch_id: str, sort_by: str = "net_repayment", order: str = "desc"):
    """返回候选清单 JSON（用于前端排序/筛选）。"""
    state = _pipeline_states.get(batch_id, {})
    candidates = state.get("result", {}).get("candidates", [])

    reverse = order == "desc"
    if sort_by in candidates[0] if candidates else []:
        candidates = sorted(candidates, key=lambda x: x.get(sort_by, 0), reverse=reverse)

    return {"candidates": candidates, "count": len(candidates)}


@router.get("/export/{batch_id}")
async def export_excel(batch_id: str):
    """导出候选清单为 Excel。"""
    import pandas as pd

    state = _pipeline_states.get(batch_id, {})
    candidates = state.get("result", {}).get("candidates", [])

    if not candidates:
        return {"error": "无数据"}

    # 展开 keywords 列表为字符串
    rows = []
    for c in candidates:
        r = {k: v for k, v in c.items() if k not in ("keywords", "json_keywords")}
        r["keywords"] = ", ".join(c["keywords"])
        r["cpc"] = f"${c['cpc']:.2f}"
        r["price"] = f"${c['price']:.2f}"
        r["fba_fee"] = f"${c['fba_fee']:.2f}"
        r["platform_fee"] = f"${c['platform_fee']:.2f}"
        r["profit_rate"] = f"{c['profit_rate']}%"
        r["net_repayment"] = f"${c['net_repayment']:.2f}"
        rows.append(r)

    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="候选清单", index=False)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=candidates_{batch_id}.xlsx"},
    )


@router.get("/aba", response_class=HTMLResponse)
async def aba_import_page(request: Request):
    """ABA 月度报表导入页。"""
    return templates.TemplateResponse(request, "aba_import.html")


@router.get("/batches", response_class=HTMLResponse)
async def list_batches(request: Request):
    """批次历史列表。"""
    db = SessionLocal()
    try:
        batches = db.query(Batch).order_by(Batch.created_at.desc()).limit(20).all()
        return templates.TemplateResponse(request, "batches.html", {
            "batches": batches,
        })
    finally:
        db.close()


# ── ABA 月度报表导入 API ──────────────────────

@router.post("/api/aba/import")
async def api_aba_import(file: UploadFile = File(...)):
    """导入 ABA 月度报表 CSV 到 aba_report 表。用 SQL INSERT OR REPLACE 批量写入。"""
    import time

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext != "csv":
        return {"error": "仅支持 CSV 格式"}

    content = await file.read()
    filepath = UPLOAD_DIR / file.filename
    filepath.write_bytes(content)

    meta = parse_metadata(str(filepath))
    domain = meta["domain"]
    report_date = meta["report_date"]

    t0 = time.time()
    df = parse_aba_excel(str(filepath))
    col_map = detect_aba_format(df)

    if not col_map.get("keyword"):
        return {"error": "未找到关键词列"}

    kw_col = col_map["keyword"]
    rank_col = col_map.get("rank")
    brand_cols = col_map.get("brand_cols", [])
    asin_cols = col_map.get("asin_cols", [])
    click_cols = sorted([c for c in df.columns if "点击份额" in c])
    conv_cols = sorted([c for c in df.columns if "转化份额" in c])
    title_cols = sorted([c for c in df.columns if "商品名称" in c])
    cat_cols = sorted([c for c in df.columns if ("类别" in c) and ("品牌" not in c)])
    date_col = next((c for c in df.columns if "报告日期" in c), None)

    # 清洗数据，构建 records 列表
    def _val(row, col):
        if col and col in row.index:
            v = str(row[col]).strip()
            return v if v not in ("0", "nan", "None") else ""
        return ""

    records = []
    for _, row in df.iterrows():
        keyword = str(row[kw_col]).strip()
        if not keyword or keyword in ("0", "nan"):
            continue

        row_date = report_date
        if date_col and str(row.get(date_col, "")).strip() not in ("0", "nan", ""):
            row_date = str(row[date_col]).strip()

        records.append({
            "keyword": keyword, "domain": domain, "report_date": row_date,
            "search_rank": int(float(str(row[rank_col]).replace(",", ""))) if rank_col and rank_col in row.index else 0,
            "brand_1": _val(row, brand_cols[0]) if len(brand_cols) > 0 else "",
            "brand_2": _val(row, brand_cols[1]) if len(brand_cols) > 1 else "",
            "brand_3": _val(row, brand_cols[2]) if len(brand_cols) > 2 else "",
            "category_1": _val(row, cat_cols[0]) if len(cat_cols) > 0 else "",
            "category_2": _val(row, cat_cols[1]) if len(cat_cols) > 1 else "",
            "category_3": _val(row, cat_cols[2]) if len(cat_cols) > 2 else "",
            "asin_1": _val(row, asin_cols[0]) if len(asin_cols) > 0 else "",
            "asin_1_title": _val(row, title_cols[0]) if len(title_cols) > 0 else "",
            "asin_1_click_share": float(row[click_cols[0]]) if len(click_cols) > 0 and click_cols[0] in row.index else 0,
            "asin_1_conversion_share": float(row[conv_cols[0]]) if len(conv_cols) > 0 and conv_cols[0] in row.index else 0,
            "asin_2": _val(row, asin_cols[1]) if len(asin_cols) > 1 else "",
            "asin_2_title": _val(row, title_cols[1]) if len(title_cols) > 1 else "",
            "asin_2_click_share": float(row[click_cols[1]]) if len(click_cols) > 1 and click_cols[1] in row.index else 0,
            "asin_2_conversion_share": float(row[conv_cols[1]]) if len(conv_cols) > 1 and conv_cols[1] in row.index else 0,
            "asin_3": _val(row, asin_cols[2]) if len(asin_cols) > 2 else "",
            "asin_3_title": _val(row, title_cols[2]) if len(title_cols) > 2 else "",
            "asin_3_click_share": float(row[click_cols[2]]) if len(click_cols) > 2 and click_cols[2] in row.index else 0,
            "asin_3_conversion_share": float(row[conv_cols[2]]) if len(conv_cols) > 2 and conv_cols[2] in row.index else 0,
        })

    # 批量 INSERT OR REPLACE（利用 SQLite 的 UPSERT）
    from sqlalchemy import text
    db = SessionLocal()
    try:
        db.execute(text("DELETE FROM aba_report WHERE domain = :d AND report_date = :rd"), {"d": domain, "rd": report_date})
        db.commit()

        db.bulk_insert_mappings(AbaReport, records)
        db.commit()

        t1 = time.time()
        return {
            "filename": file.filename,
            "domain": domain,
            "report_date": report_date,
            "total_rows": len(records),
            "elapsed_sec": round(t1 - t0, 1),
        }
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


@router.get("/api/aba/stats")
async def api_aba_stats():
    """查看 aba_report 表统计信息。"""
    from sqlalchemy import func
    db = SessionLocal()
    try:
        total = db.query(func.count(AbaReport.id)).scalar()
        domains = db.query(AbaReport.domain, func.count(AbaReport.id)).group_by(AbaReport.domain).all()
        dates = db.query(AbaReport.report_date, func.count(AbaReport.id)).group_by(AbaReport.report_date).all()
        return {
            "total": total,
            "by_domain": {d[0]: d[1] for d in domains},
            "by_date": {d[0]: d[1] for d in dates},
        }
    finally:
        db.close()


@router.get("/api/aba/dates")
async def api_aba_dates():
    """返回已导入的 ABA 数据月份列表，供首页下拉框。"""
    from sqlalchemy import func
    db = SessionLocal()
    try:
        rows = db.query(
            AbaReport.domain, AbaReport.report_date, func.count(AbaReport.id)
        ).group_by(AbaReport.domain, AbaReport.report_date).order_by(
            AbaReport.report_date.desc()
        ).all()
        return {
            "dates": [
                {"value": f"{r[0]}|{r[1]}", "label": f"{'🇺🇸 US' if r[0]=='US' else '🇨🇦 CA' if r[0]=='CA' else r[0]} — {r[1]} ({r[2]:,} 条)"}
                for r in rows
            ]
        }
    finally:
        db.close()


# ── 全自动管线 ──────────────────────────────

@router.post("/api/pipeline/auto-run")
async def api_auto_run(
    file: UploadFile = File(...),
    domain: str = Form(""),
    conv_index_min: float = Form(1.0),
    share_max: float = Form(50.0),
    ship_cbm_rate: float = Form(120),
    ship_handling: float = Form(0.8),
    data_expiry_days: int = Form(30),
):
    """一键全自动：上传 ABA CSV → 自动跑卖家精灵 → 入库 → 管线筛选。"""
    from .pipeline import run_auto_pipeline

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls"):
        return {"error": "仅支持 CSV / Excel 格式"}

    content = await file.read()
    filepath = UPLOAD_DIR / f"aba_auto_{os.urandom(4).hex()}.{ext}"
    filepath.write_bytes(content)

    # 优先用前端传的站点，否则从文件名检测，最后回退 CA
    domain_str = domain.strip().upper() if domain else ""
    valid_domains = {"US", "UK", "DE", "JP", "CA"}
    if domain_str not in valid_domains:
        domain_str = "CA"
        basename = file.filename.upper()
        for d in valid_domains:
            if basename.startswith(d + "_"):
                domain_str = d
                break

    config = {
        "conv_index_min": conv_index_min,
        "share_max": share_max,
        "data_expiry_days": data_expiry_days,
        "ship_cbm_rate": ship_cbm_rate,
        "ship_handling": ship_handling,
    }
    batch_id = f"auto_{os.urandom(4).hex()}"

    progress = PipelineProgress()
    _pipeline_states[batch_id] = {"status": "running", "progress": progress}

    def _auto_thread():
        try:
            result = run_auto_pipeline(
                aba_filepath=str(filepath),
                domain_str=domain_str,
                config=config,
                progress=progress,
            )
            _pipeline_states[batch_id] = {"status": "done", "result": result, "progress": progress}
        except Exception as e:
            _pipeline_states[batch_id] = {"status": "failed", "error": str(e), "progress": progress}

    thread = threading.Thread(target=_auto_thread, daemon=True)
    thread.start()

    return {"batch_id": batch_id, "domain": domain_str}


# ── Pipeline API（给新流程页用） ───────────────

@router.post("/api/pipeline/run")
async def api_run_pipeline(request: Request):
    """启动 pipeline，返回 batch_id。"""
    body = await request.json()
    date_key = body.get("date_key", "")
    config = body.get("config", {})
    data_source = body.get("data_source", "sellersprite")
    mock = body.get("mock", data_source != "sellersprite")

    if "|" not in date_key:
        return {"error": "请选择数据源"}

    domain_str, report_date = date_key.split("|", 1)
    domain_map = {"US": 1, "UK": 2, "DE": 3, "JP": 6, "CA": 7}
    domain_int = domain_map.get(domain_str, 7)

    batch_id = f"batch_{os.urandom(4).hex()}"

    # 导出 aba_report 数据为 CSV 给 pipeline 用
    db = SessionLocal()
    try:
        rows = db.query(AbaReport).filter(
            AbaReport.domain == domain_str,
            AbaReport.report_date == report_date,
        ).all()
        if not rows:
            return {"error": f"未找到 {domain_str}/{report_date} 的数据，请先导入"}

        import pandas as pd
        data = [{
            "搜索频率排名": r.search_rank,
            "搜索词": r.keyword,
            "点击量最高的品牌 #1": r.brand_1 or "",
            "点击量最高的品牌 #2": r.brand_2 or "",
            "点击量最高的品牌 #3": r.brand_3 or "",
            "点击量最高的类别 #1": r.category_1 or "",
            "点击量最高的类别 #2": r.category_2 or "",
            "点击量最高的类别 #3": r.category_3 or "",
            "点击量最高的商品 #1：ASIN": r.asin_1 or "",
            "点击量最高的商品 #1：商品名称": r.asin_1_title or "",
            "点击量最高的商品 #1：点击份额": r.asin_1_click_share or 0,
            "点击量最高的商品 #1：转化份额": r.asin_1_conversion_share or 0,
            "点击量最高的商品 #2：ASIN": r.asin_2 or "",
            "点击量最高的商品 #2：商品名称": r.asin_2_title or "",
            "点击量最高的商品 #2：点击份额": r.asin_2_click_share or 0,
            "点击量最高的商品 #2：转化份额": r.asin_2_conversion_share or 0,
            "点击量最高的商品 #3：ASIN": r.asin_3 or "",
            "点击量最高的商品 #3：商品名称": r.asin_3_title or "",
            "点击量最高的商品 #3：点击份额": r.asin_3_click_share or 0,
            "点击量最高的商品 #3：转化份额": r.asin_3_conversion_share or 0,
            "报告日期": r.report_date,
        } for r in rows]

        filepath = UPLOAD_DIR / f"aba_{batch_id}.csv"
        pd.DataFrame(data).to_csv(str(filepath), index=False, encoding="utf-8-sig")
    finally:
        db.close()

    thread = threading.Thread(
        target=_pipeline_thread,
        args=(batch_id, str(filepath), domain_int, config, mock, data_source),
        daemon=True,
    )
    thread.start()

    return {"batch_id": batch_id}


@router.get("/api/pipeline/progress/{batch_id}")
async def api_pipeline_progress(batch_id: str):
    """返回 pipeline 进度 JSON（给前端轮询）。"""
    state = _pipeline_states.get(batch_id, {})
    if not state:
        return {"status": "not_found", "percent": 0, "message": "批次未找到"}

    if state["status"] == "running":
        p = state["progress"]
        return {
            "status": "running",
            "percent": p.percent,
            "message": p.message,
            "logs": p.logs[-15:],
            "funnel": _build_funnel(state),
        }
    elif state["status"] == "done":
        result = state["result"]
        return {
            "status": "done",
            "percent": 100,
            "message": f"完成! {result['candidate_count']} 个候选 ASIN",
            "logs": state.get("progress", PipelineProgress()).logs[-15:],
            "funnel": _build_funnel(state),
            "candidates": result.get("candidates", []),
            "summary": result.get("summary", {}),
        }
    else:
        return {
            "status": "failed",
            "percent": 0,
            "message": f"失败: {state.get('error', '未知错误')}",
            "error": state.get("error"),
            "logs": state.get("progress", PipelineProgress()).logs[-15:],
        }


# ── 卖家精灵数据导入 ──────────────────────────

@router.post("/api/sellersprite/import/kcr")
async def api_import_kcr(file: UploadFile = File(...), domain: str = Form("CA")):
    """上传 KCR 导出 Excel，解析入库。"""
    from .sellersprite_import import parse_kcr_excel, import_kcr_to_db

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        return {"error": "仅支持 Excel 格式"}

    content = await file.read()
    filepath = UPLOAD_DIR / f"kcr_{os.urandom(4).hex()}.{ext}"
    filepath.write_bytes(content)

    try:
        records = parse_kcr_excel(str(filepath))
        if not records:
            return {"error": "解析结果为空，请检查文件格式"}
        batch_id = f"kcr_{os.urandom(4).hex()}"
        count = import_kcr_to_db(records, domain, batch_id)
        db = SessionLocal()
        try:
            db.add(Batch(batch_id=batch_id, domain={"US": 1, "CA": 7, "UK": 2}.get(domain, 7), status="kcr_imported", total_aba_rows=count))
            db.commit()
        finally:
            db.close()
        return {"success": True, "total_rows": len(records), "imported": count, "batch_id": batch_id, "domain": domain}
    except Exception as e:
        return {"error": str(e)}


@router.post("/api/sellersprite/import/product")
async def api_import_product(file: UploadFile = File(...), domain: str = Form("CA")):
    """上传产品库导出 Excel，解析入库。"""
    from .sellersprite_import import parse_product_excel, import_product_to_db

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        return {"error": "仅支持 Excel 格式"}

    content = await file.read()
    filepath = UPLOAD_DIR / f"product_{os.urandom(4).hex()}.{ext}"
    filepath.write_bytes(content)

    try:
        records = parse_product_excel(str(filepath))
        if not records:
            return {"error": "解析结果为空，请检查文件格式"}
        batch_id = f"prod_{os.urandom(4).hex()}"
        count = import_product_to_db(records, domain, batch_id)
        db = SessionLocal()
        try:
            db.add(Batch(batch_id=batch_id, domain={"US": 1, "CA": 7, "UK": 2}.get(domain, 7), status="product_imported", total_aba_rows=count))
            db.commit()
        finally:
            db.close()
        return {"success": True, "total_rows": len(records), "imported": count, "batch_id": batch_id, "domain": domain}
    except Exception as e:
        return {"error": str(e)}


@router.get("/api/sellersprite/stats")
async def api_sellersprite_stats(domain: str = Query("CA")):
    """卖家精灵已导入数据概览。"""
    db = SessionLocal()
    try:
        kw_count = db.query(SellerspriteKeyword).filter(SellerspriteKeyword.domain == domain).count()
        prod_count = db.query(SellerspriteProduct).filter(SellerspriteProduct.domain == domain).count()

        # 有数据的 keyword 和 product batch 列表
        kw_batches = db.query(SellerspriteKeyword.batch_id).filter(SellerspriteKeyword.domain == domain).distinct().all()
        prod_batches = db.query(SellerspriteProduct.batch_id).filter(SellerspriteProduct.domain == domain).distinct().all()

        return {
            "domain": domain,
            "keyword_count": kw_count,
            "product_count": prod_count,
            "keyword_batches": [b[0] for b in kw_batches],
            "product_batches": [b[0] for b in prod_batches],
        }
    finally:
        db.close()


# ── 导出关键词/ASIN 列表（给独立脚本用） ──────

@router.get("/api/pipeline/export/keywords/{date_key}")
async def export_keywords(date_key: str):
    """从 aba_report 导出去重关键词列表（txt 格式，给 kcr_batch.py 用）。"""
    if "|" not in date_key:
        return {"error": "无效的 date_key"}

    domain_str, report_date = date_key.split("|", 1)
    db = SessionLocal()
    try:
        rows = db.query(AbaReport.keyword).filter(
            AbaReport.domain == domain_str,
            AbaReport.report_date == report_date,
        ).distinct().order_by(AbaReport.keyword).all()

        keywords = [r[0] for r in rows if r[0]]
        content = "\n".join(keywords)

        return StreamingResponse(
            iter([content]),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=keywords_{domain_str}_{report_date}.txt"},
        )
    finally:
        db.close()


@router.get("/api/pipeline/export/asins/{date_key}")
async def export_asins(date_key: str):
    """从 aba_report 导出去重 ASIN 列表（txt 格式，给 asin_batch.py 用）。"""
    if "|" not in date_key:
        return {"error": "无效的 date_key"}

    domain_str, report_date = date_key.split("|", 1)
    db = SessionLocal()
    try:
        rows = db.query(AbaReport).filter(
            AbaReport.domain == domain_str,
            AbaReport.report_date == report_date,
        ).all()

        asins = set()
        for r in rows:
            for a in [r.asin_1, r.asin_2, r.asin_3]:
                if a and a.startswith("B0"):
                    asins.add(a)

        content = "\n".join(sorted(asins))

        return StreamingResponse(
            iter([content]),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=asins_{domain_str}_{report_date}.txt"},
        )
    finally:
        db.close()


def _build_funnel(state: dict) -> list:
    """从 pipeline 状态构建漏斗数据。"""
    result = state.get("result", {})
    summary = result.get("summary", {})
    p = state.get("progress")
    logs = p.logs if p else []

    total = summary.get("total_aba_rows", 0)
    final = summary.get("final_candidates", 0)
    step3 = summary.get("step3_passed", 0)

    # 从 [FUNNEL] 日志行解析
    conv_removed = 0
    brand_removed = 0
    share_removed = 0
    passed = 0
    import re
    for log in logs:
        if "[FUNNEL]" in log:
            m = re.search(r'total=(\d+)', log)
            if m: total = int(m.group(1))
            m = re.search(r'conv_removed=(\d+)', log)
            if m: conv_removed = int(m.group(1))
            m = re.search(r'brand_removed=(\d+)', log)
            if m: brand_removed = int(m.group(1))
            m = re.search(r'share_removed=(\d+)', log)
            if m: share_removed = int(m.group(1))
            m = re.search(r'passed=(\d+)', log)
            if m: passed = int(m.group(1))

    after_conv = total - conv_removed
    after_brand = after_conv - brand_removed
    after_share = after_brand - share_removed
    biz_removed = max(0, after_share - step3) if step3 > 0 else 0

    status = state.get("status", "") if isinstance(state, dict) else ""
    is_running = status == "running"
    is_done = status == "done"

    def _status(has_count):
        if is_done: return "done"
        if is_running and has_count: return "done"
        if is_running: return "running"
        return "pending"

    return [
        {"label": "原始数据", "count": total, "removed": 0, "status": _status(total > 0)},
        {"label": "转化系数过滤", "count": after_conv, "removed": conv_removed, "status": _status(passed > 0 or after_conv > 0)},
        {"label": "品牌词过滤", "count": after_brand, "removed": brand_removed, "status": _status(after_brand > 0)},
        {"label": "集中度过滤", "count": after_share, "removed": share_removed, "status": _status(after_share > 0)},
        {"label": "商机探测器", "count": step3, "removed": biz_removed, "status": _status(step3 > 0)},
        {"label": "候选产品", "count": final, "removed": 0, "status": _status(final > 0)},
    ]


# ── 凭据管理 API ─────────────────────────────

CREDENTIALS_FILE = Path.home() / ".sellersprite-credentials"


@router.get("/api/credentials")
async def api_read_credentials():
    """读取当前凭据（密码脱敏）。"""
    try:
        data = json.loads(CREDENTIALS_FILE.read_text(encoding="utf-8"))
        email = data.get("email", "")
        pwd_len = len(data.get("password", ""))
        return {"email": email, "password_masked": "*" * pwd_len, "has_credentials": True}
    except Exception:
        return {"email": "", "password_masked": "", "has_credentials": False}


@router.post("/api/credentials")
async def api_save_credentials(request: Request):
    """保存凭据到 ~/.sellersprite-credentials。"""
    body = await request.json()
    email = body.get("email", "").strip()
    password = body.get("password", "").strip()
    if not email or not password:
        return {"error": "邮箱和密码不能为空"}
    data = {"email": email, "password": password}
    CREDENTIALS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Niche 分析 API ─────────────────────────────

@router.get("/niche/{batch_id}", response_class=HTMLResponse)
async def niche_list_page(request: Request, batch_id: str):
    """Niche 列表页。"""
    return templates.TemplateResponse(request, "niche_list.html", {"batch_id": batch_id})


@router.get("/niche/{batch_id}/{niche_name}", response_class=HTMLResponse)
async def niche_detail_page(request: Request, batch_id: str, niche_name: str):
    """Niche 详情页。"""
    return templates.TemplateResponse(request, "niche_detail.html", {
        "batch_id": batch_id,
        "niche_name": niche_name,
    })


@router.get("/api/niche/{batch_id}")
async def api_niche_list(batch_id: str):
    """返回 niche 列表（含摘要数据）。"""
    from .niche_analyzer import form_niches
    from .models import SessionLocal, Candidate

    db = SessionLocal()
    try:
        rows = db.query(Candidate).filter(
            Candidate.batch_id == batch_id, Candidate.price > 0
        ).all()
        candidates = []
        for c in rows:
            net_rate = c.net_repayment / c.price if c.price > 0 else 0
            if net_rate < 0.5:
                continue  # 淘汰，DB保留
            candidates.append({
                "asin": c.asin,
                "keywords": json.loads(c.keywords) if c.keywords else [],
                "price": c.price,
                "brand": c.brand,
                "net_repayment": c.net_repayment,
            })
        if not candidates:
            return {"niche_count": 0, "niches": []}

        niches = form_niches(candidates)
        # 只返回 ASIN ≥ 2 的
        result = []
        for name, niche in sorted(niches.items(), key=lambda x: -len(x[1]["asins"])):
            if len(niche["asins"]) < 2:
                continue
            asin_list = sorted(niche["asins"])
            prices = [c["price"] for c in candidates if c["asin"] in niche["asins"]]
            brands = list({c["brand"] for c in candidates if c["asin"] in niche["asins"] and c["brand"]})
            result.append({
                "name": name,
                "core_count": len(asin_list),
                "keyword_count": len(niche["keywords"]),
                "core_asins": asin_list,
                "keywords": sorted(niche["keywords"]),
                "price_min": min(prices) if prices else 0,
                "price_max": max(prices) if prices else 0,
                "brands": brands[:5],
            })
        return {"niche_count": len(result), "niches": result}
    finally:
        db.close()


@router.get("/api/niche/{batch_id}/{niche_name}")
async def api_niche_detail(batch_id: str, niche_name: str):
    """返回单个 niche 的详细数据（产品 + 竞争矩阵）。"""
    from .niche_analyzer import run_niche_analysis
    result = run_niche_analysis(batch_id)
    for n in result.get("niches", []):
        if n["name"] == niche_name:
            return n
    return {"error": "niche 未找到"}


# ── 数据库维护 API ─────────────────────────────

@router.get("/api/maintenance/stats")
async def api_maintenance_stats():
    """返回数据库各表行数和大小。"""
    from sqlalchemy import text as sa_text
    import os
    from .models import DB_PATH

    db = SessionLocal()
    try:
        tables = db.execute(sa_text(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        )).fetchall()

        stats = []
        for (tname,) in tables:
            count = db.execute(sa_text(f"SELECT COUNT(*) FROM \"{tname}\"")).scalar()
            stats.append({"table": tname, "rows": count})

        db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
        return {
            "db_size_mb": round(db_size / 1024 / 1024, 2),
            "tables": stats,
        }
    finally:
        db.close()


@router.post("/api/maintenance/cleanup")
async def api_maintenance_cleanup(keep: int = 3):
    """手动清理旧批次数据。"""
    from .pipeline import _cleanup_old_batches
    from sqlalchemy import text as sa_text
    import os
    from .models import DB_PATH

    size_before = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    _cleanup_old_batches(keep=keep)

    # VACUUM 回收空间
    db = SessionLocal()
    try:
        db.execute(sa_text("VACUUM"))
        db.commit()
    finally:
        db.close()

    size_after = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    return {
        "ok": True,
        "keep_batches": keep,
        "size_before_mb": round(size_before / 1024 / 1024, 2),
        "size_after_mb": round(size_after / 1024 / 1024, 2),
    }


# ════════════════════════════════════════════════════════════════
# 赛道探索器 API
# ════════════════════════════════════════════════════════════════
#
# GET  /explorer              — 探索器主页面
# POST /api/explorer/upload    — 上传文件 → 关键词列表（固定模板）
# ════════════════════════════════════════════════════════════════

_MAX_FILE_MB = 10

# 固定列模板：关键词 | 月搜索量 | CPC | 月销量
# 英文：keyword | search_volume | cpc | sales
_TEMPLATE_COLS = {
    "keyword":  ["关键词", "keyword", "搜索词", "search_term", "searchterm", "query"],
    "流量":     ["点击量", "月点击量", "clicks", "click", "点击"],
    "cpc":      ["cpc", "ppc", "PPC竞价", "竞价"],
    "sales":    ["月购买量", "月销量", "购买量", "销量", "sales", "salesvolume"],
}


@router.get("/explorer", response_class=HTMLResponse)
async def explorer_page(request: Request):
    return templates.TemplateResponse(request, "explorer.html")


_TEMPLATE_COLS = {
    "keyword":  ["关键词", "keyword", "搜索词", "search_term", "searchterm", "query"],
    "流量":     ["点击量", "月点击量", "clicks", "click", "点击"],
    "cpc":      ["cpc", "ppc", "PPC竞价", "竞价"],
    "sales":    ["月购买量", "月销量", "购买量", "销量", "sales", "salesvolume"],
}


@router.post("/api/explorer/upload")
async def explorer_upload(file: UploadFile | None = File(None)):
    if not file or not file.filename:
        return {"ok": False, "error": "请上传文件 (CSV/Excel)"}
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('csv', 'xlsx', 'xls'):
        return {"ok": False, "error": f"不支持 .{ext}，请上传 CSV 或 Excel"}
    filepath = UPLOAD_DIR / f"explorer_{file.filename}"
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        return {"ok": False, "error": "文件过大（>10MB）"}
    filepath.write_bytes(content)

    rows_raw = []
    if ext == 'csv':
        import csv as _csv
        with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
            for r in _csv.DictReader(f):
                rows_raw.append(r)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        sheet = wb.active
        it = sheet.iter_rows()
        hdr = next(it, None)
        if not hdr: return {"ok": False, "error": "文件为空"}
        headers = [str(c.value or '').strip() for c in hdr]
        for row in it:
            rows_raw.append(dict(zip(headers, [str(c.value or '').strip() for c in row])))
        wb.close()

    if not rows_raw: return {"ok": False, "error": "文件中无数据行"}

    valid_headers = [str(h).strip() for h in rows_raw[0].keys() if h is not None and str(h).strip()]
    mapping = {}
    for field, candidates in _TEMPLATE_COLS.items():
        for h in valid_headers:
            hl = h.lower().replace(' ', '').replace('_', '')
            if any(c.lower().replace(' ', '').replace('_', '') == hl for c in candidates):
                mapping[field] = h; break
        if not mapping.get(field):
            for h in valid_headers:
                hl = h.lower().replace(' ', '').replace('_', '')
                if any(c.lower().replace(' ', '').replace('_', '') in hl for c in candidates):
                    mapping[field] = h; break
    if not mapping.get("keyword"):
        return {"ok": False, "error": f"找不到关键词列。支持: {', '.join(_TEMPLATE_COLS['keyword'])}"}

    def _num(v):
        try: return float(str(v).replace('$','').replace('%','').replace(',','').strip())
        except: return 0

    keywords = []
    for r in rows_raw:
        kw = str(r.get(mapping["keyword"], '')).strip()
        if not kw or len(kw) < 2: continue
        clicks = int(_num(r.get(mapping.get("流量", ""), 0)))
        cpc = _num(r.get(mapping.get("cpc", ""), 0))
        sales = int(_num(r.get(mapping.get("sales", ""), 0)))
        if clicks <= 0 and sales <= 0: continue
        conv = round(sales / clicks, 4) if clicks > 0 else 0
        keywords.append({"kw": kw, "流量": clicks, "cpc": cpc, "sales": sales, "conv": conv})

    if not keywords: return {"ok": False, "error": "未解析到有效关键词"}
    return {"ok": True, "keywords": keywords, "mapping": {k: v for k, v in mapping.items()}}


# ═══════════════════════════════════════════════════════
# 词根赛道 Dashboard 路由
# ═══════════════════════════════════════════════════════

@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(request: Request):
    return templates.TemplateResponse(request, "dashboard.html")


@router.post("/api/dashboard/upload/keywords")
async def dashboard_upload_keywords(file: UploadFile = File(...), domain: str = Form("US")):
    """上传关键词挖掘 Excel，解析入库。自动从文件名提取数据月份。"""
    from .keyword_mining_parser import parse_keyword_mining, parse_unique_words, import_keyword_mining_to_db, import_unique_words_to_db, extract_period_from_filename

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        return {"error": "仅支持 Excel 格式"}

    filepath = UPLOAD_DIR / f"kw_mining_{os.urandom(4).hex()}.{ext}"
    filepath.write_bytes(await file.read())

    try:
        data_period = extract_period_from_filename(file.filename)
        kw_records = parse_keyword_mining(str(filepath))
        if not kw_records:
            return {"error": "解析结果为空，请检查文件格式"}
        batch_label = f"{file.filename.rsplit('.', 1)[0]}-{os.urandom(2).hex()}"
        kw_count = import_keyword_mining_to_db(kw_records, domain, batch_label, data_period)

        uw_records = parse_unique_words(str(filepath))
        uw_count = import_unique_words_to_db(uw_records, domain, batch_label) if uw_records else 0

        return {
            "success": True,
            "batch_label": batch_label,
            "keyword_count": kw_count,
            "word_root_count": uw_count,
            "domain": domain,
            "data_period": data_period,
        }
    except Exception as e:
        return {"error": str(e)}


@router.post("/api/dashboard/upload/products")
async def dashboard_upload_products(file: UploadFile = File(...), domain: str = Form("US"), start_month: str = Form("")):
    """上传产品库导出 Excel，解析入库。按文件名序号 + 起始月计算 data_period。"""
    from .sellersprite_import import parse_product_excel, import_product_to_db, parse_product_sequence, compute_period

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        return {"error": "仅支持 Excel 格式"}

    filepath = UPLOAD_DIR / f"product_{os.urandom(4).hex()}.{ext}"
    filepath.write_bytes(await file.read())

    try:
        records = parse_product_excel(str(filepath))
        if not records:
            return {"error": "解析结果为空，请检查文件格式"}
        seq = parse_product_sequence(file.filename)
        data_period = compute_period(start_month, seq) if start_month else ""
        batch_label = f"{file.filename.rsplit('.', 1)[0]}-{os.urandom(2).hex()}"
        count = import_product_to_db(records, domain, batch_label, data_period=data_period)

        return {
            "success": True,
            "batch_label": batch_label,
            "product_count": count,
            "domain": domain,
            "data_period": data_period,
            "sequence": seq,
        }
    except Exception as e:
        return {"error": str(e)}


@router.get("/api/dashboard/batches")
async def dashboard_batches(domain: str = Query("US")):
    """列出已导入的批次。"""
    db = SessionLocal()
    try:
        kw_batches = db.query(SellerspriteKeyword.batch_id).filter(
            SellerspriteKeyword.domain == domain
        ).distinct().all()
        prod_batches = db.query(SellerspriteProduct.batch_id).filter(
            SellerspriteProduct.domain == domain
        ).distinct().all()
        return {
            "keyword_batches": [b[0] for b in kw_batches if b[0]],
            "product_batches": [b[0] for b in prod_batches if b[0]],
        }
    finally:
        db.close()


@router.get("/api/dashboard/word-roots")
async def dashboard_word_roots(keyword_batch_label: str = Query(...), domain: str = Query("US")):
    """获取某批次的所有词根。"""
    db = SessionLocal()
    try:
        roots = db.query(WordRoot).filter(
            WordRoot.batch_label == keyword_batch_label,
            WordRoot.domain == domain,
        ).order_by(WordRoot.frequency.desc()).all()
        return {
            "roots": [{"word": r.word, "frequency": r.frequency} for r in roots],
            "total": len(roots),
        }
    finally:
        db.close()


@router.get("/api/dashboard/asins")
async def dashboard_asins(keyword_batch_label: str = Query(...), domain: str = Query("US")):
    """提取某批次关键词的所有 TOP10 ASIN，去重返回。"""
    db = SessionLocal()
    try:
        rows = db.query(SellerspriteKeyword.top10_asins).filter(
            SellerspriteKeyword.batch_id == keyword_batch_label,
            SellerspriteKeyword.domain == domain,
        ).all()
        asin_set = set()
        for (top10,) in rows:
            if top10:
                for a in top10.split(","):
                    a = a.strip()
                    if a.startswith("B0"):
                        asin_set.add(a)
        asins = sorted(asin_set)
        return {"asins": asins, "count": len(asins)}
    finally:
        db.close()


@router.get("/api/dashboard/periods")
async def dashboard_periods(domain: str = Query("US")):
    """获取所有已导入数据的月份列表。"""
    db = SessionLocal()
    try:
        kw_periods = db.query(SellerspriteKeyword.data_period).filter(
            SellerspriteKeyword.domain == domain,
            SellerspriteKeyword.data_period != "",
        ).distinct().all()
        prod_periods = db.query(SellerspriteProduct.data_period).filter(
            SellerspriteProduct.domain == domain,
            SellerspriteProduct.data_period != "",
        ).distinct().all()
        all_periods = sorted(set(p[0] for p in kw_periods + prod_periods if p[0]))
        return {"periods": all_periods}
    finally:
        db.close()


@router.post("/api/dashboard/filter")
async def dashboard_filter(body: dict):
    """按词根组合筛选关键词和 ASIN。可选按 data_period 过滤。"""
    domain = body.get("domain", "US")
    root_words = body.get("root_words", [])
    data_period = body.get("data_period", "")

    db = SessionLocal()
    try:
        # 加载关键词（按period过滤或全部）
        kw_query = db.query(SellerspriteKeyword).filter(
            SellerspriteKeyword.domain == domain,
        )
        if data_period:
            kw_query = kw_query.filter(SellerspriteKeyword.data_period == data_period)
        kw_rows = kw_query.all()

        # 词根交集筛选
        stop_words = _get_stop_words()
        matched_keywords = []
        for row in kw_rows:
            if root_words:
                kw_lower = row.keyword.lower()
                words = set(w for w in re.split(r"[^a-z0-9]+", kw_lower) if len(w) >= 2 and w not in stop_words)
                if not all(r in words for r in root_words):
                    continue
            matched_keywords.append(row)

        # 从匹配关键词提取 ASIN + 加载产品
        matched_asins_set = set()
        for row in matched_keywords:
            if row.top10_asins:
                for a in row.top10_asins.split(","):
                    a = a.strip()
                    if a.startswith("B0"):
                        matched_asins_set.add(a)

        products = {}
        if matched_asins_set:
            prod_query = db.query(SellerspriteProduct).filter(
                SellerspriteProduct.domain == domain,
                SellerspriteProduct.asin.in_(matched_asins_set),
            )
            if data_period:
                prod_query = prod_query.filter(SellerspriteProduct.data_period == data_period)
            prod_rows = prod_query.all()
            products = {p.asin: p for p in prod_rows}

        # 构建关键词级数据
        kw_data = []
        for row in matched_keywords:
            raw = json.loads(row.raw_response) if row.raw_response else {}
            kw_data.append({
                "keyword": row.keyword,
                "keyword_cn": row.keyword_cn or raw.get("keyword_cn", ""),
                "search_volume": row.search_volume or 0,
                "purchases": row.purchases_90d or 0,
                "purchase_rate": round((row.search_conversion_rate or 0), 1),
                "cpc": round(row.cpc_recommended or 0, 2),
                "avg_price": round(row.avg_price or 0, 2),
                "click_share": round((row.click_share or 0), 1),
                "conv_share": round((row.conv_share or 0), 1),
                "product_count": raw.get("product_count", 0),
                "review_count": raw.get("review_count", 0),
                "rating": raw.get("rating", 0),
                "category": raw.get("category", ""),
                "impressions": raw.get("impressions", 0),
                "clicks": raw.get("clicks", 0),
            })

        # 构建 ASIN 级数据
        asin_data = []
        for asin in sorted(matched_asins_set):
            p = products.get(asin)
            raw = json.loads(p.raw_response) if p and p.raw_response else {}
            asin_data.append({
                "asin": asin,
                "title": p.title if p else "",
                "brand": p.brand if p else "",
                "price": round(p.price, 2) if p and p.price else 0,
                "monthly_sales": p.monthly_sales if p else 0,
                "monthly_revenue": round(p.monthly_revenue, 2) if p and p.monthly_revenue else 0,
                "bsr": p.main_bsr if p else 0,
                "ratings_count": p.ratings_count if p else 0,
                "ratings": round(p.ratings, 1) if p and p.ratings else 0,
                "online_days": p.online_days if p else 0,
                "is_fba": bool(p.is_fba) if p else False,
                "category": p.main_category if p else "",
                "seller_count": p.seller_count if p else 0,
            })

        # 汇总
        summary = {
            "keyword_count": len(kw_data),
            "asin_count": len(asin_data),
            "total_sv": sum(k["search_volume"] for k in kw_data),
            "total_purchases": sum(k["purchases"] for k in kw_data),
            "avg_cpc": round(sum(k["cpc"] for k in kw_data) / max(len(kw_data), 1), 2),
            "avg_price": round(sum(k["avg_price"] for k in kw_data) / max(len(kw_data), 1), 2),
            "total_monthly_sales": sum(a["monthly_sales"] for a in asin_data),
            "total_monthly_revenue": round(sum(a["monthly_revenue"] for a in asin_data), 2),
        }

        # 可用月份
        kw_periods = db.query(SellerspriteKeyword.data_period).filter(
            SellerspriteKeyword.domain == domain, SellerspriteKeyword.data_period != ""
        ).distinct().all()
        prod_periods = db.query(SellerspriteProduct.data_period).filter(
            SellerspriteProduct.domain == domain, SellerspriteProduct.data_period != ""
        ).distinct().all()
        all_periods = sorted(set(p[0] for p in kw_periods + prod_periods if p[0]))

        return {"keywords": kw_data, "asins": asin_data, "summary": summary, "periods": all_periods}
    finally:
        db.close()


@router.post("/api/dashboard/niches")
async def dashboard_save_niche(body: dict):
    """保存赛道。"""
    db = SessionLocal()
    try:
        nt = NicheTrack(
            name=body.get("name", ""),
            keyword_batch_label=body.get("keyword_batch_label", ""),
            product_batch_label=body.get("product_batch_label", ""),
            domain=body.get("domain", "US"),
            root_words=json.dumps(body.get("root_words", []), ensure_ascii=False),
            keyword_count=body.get("keyword_count", 0),
            asin_count=body.get("asin_count", 0),
            stats_snapshot=json.dumps(body.get("stats_snapshot", {}), ensure_ascii=False),
        )
        db.add(nt)
        db.commit()
        db.refresh(nt)
        return {"success": True, "id": nt.id}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


@router.get("/api/dashboard/niches")
async def dashboard_list_niches(keyword_batch_label: str = Query(...)):
    """列出某批次下的所有已保存赛道。"""
    db = SessionLocal()
    try:
        rows = db.query(NicheTrack).filter(
            NicheTrack.keyword_batch_label == keyword_batch_label,
        ).order_by(NicheTrack.created_at.desc()).all()
        return {
            "niches": [{
                "id": r.id,
                "name": r.name,
                "root_words": json.loads(r.root_words) if r.root_words else [],
                "keyword_count": r.keyword_count,
                "asin_count": r.asin_count,
                "stats_snapshot": json.loads(r.stats_snapshot) if r.stats_snapshot else {},
                "created_at": r.created_at.isoformat() if r.created_at else "",
            } for r in rows],
        }
    finally:
        db.close()


@router.get("/api/dashboard/niche/{niche_id}")
async def dashboard_get_niche(niche_id: int):
    """获取单个赛道详情。"""
    db = SessionLocal()
    try:
        r = db.query(NicheTrack).filter(NicheTrack.id == niche_id).first()
        if not r:
            return {"error": "赛道不存在"}
        return {
            "id": r.id,
            "name": r.name,
            "keyword_batch_label": r.keyword_batch_label,
            "product_batch_label": r.product_batch_label,
            "domain": r.domain,
            "root_words": json.loads(r.root_words) if r.root_words else [],
            "keyword_count": r.keyword_count,
            "asin_count": r.asin_count,
            "stats_snapshot": json.loads(r.stats_snapshot) if r.stats_snapshot else {},
            "created_at": r.created_at.isoformat() if r.created_at else "",
        }
    finally:
        db.close()


@router.delete("/api/dashboard/niche/{niche_id}")
async def dashboard_delete_niche(niche_id: int):
    """删除赛道。"""
    db = SessionLocal()
    try:
        r = db.query(NicheTrack).filter(NicheTrack.id == niche_id).first()
        if not r:
            return {"error": "赛道不存在"}
        db.delete(r)
        db.commit()
        return {"success": True}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


# ═══════════════════════════════════════════════════════
# 项目 CRUD
# ═══════════════════════════════════════════════════════

@router.post("/api/projects")
async def api_create_project(request: Request):
    """创建项目。"""
    body = await request.json()
    name = body.get("name", "").strip()
    if not name:
        return {"error": "项目名称不能为空"}
    db = SessionLocal()
    try:
        p = Project(
            name=name,
            description=body.get("description", ""),
            domain=body.get("domain", "US"),
            keyword_batch_label=body.get("keyword_batch_label", ""),
            product_batch_label=body.get("product_batch_label", ""),
        )
        db.add(p)
        db.commit()
        db.refresh(p)
        return {"success": True, "id": p.id, "name": p.name}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


@router.get("/api/projects")
async def api_list_projects():
    """列出所有项目，含摘要数据。"""
    db = SessionLocal()
    try:
        projects = db.query(Project).order_by(Project.created_at.desc()).all()
        result = []
        for p in projects:
            kw_count = db.query(SellerspriteKeyword).filter(
                SellerspriteKeyword.batch_id == p.keyword_batch_label
            ).count() if p.keyword_batch_label else 0
            prod_count = db.query(SellerspriteProduct).filter(
                SellerspriteProduct.batch_id == p.product_batch_label
            ).count() if p.product_batch_label else 0
            result.append({
                "id": p.id,
                "name": p.name,
                "description": p.description,
                "domain": p.domain,
                "keyword_batch_label": p.keyword_batch_label,
                "product_batch_label": p.product_batch_label,
                "keyword_count": kw_count,
                "product_count": prod_count,
                "created_at": p.created_at.isoformat() if p.created_at else "",
            })
        return {"projects": result}
    finally:
        db.close()


@router.get("/api/projects/{project_id}")
async def api_get_project(project_id: int):
    """获取单个项目详情。"""
    db = SessionLocal()
    try:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            return {"error": "项目不存在"}
        return {
            "id": p.id, "name": p.name, "description": p.description,
            "domain": p.domain, "keyword_batch_label": p.keyword_batch_label,
            "product_batch_label": p.product_batch_label,
            "created_at": p.created_at.isoformat() if p.created_at else "",
        }
    finally:
        db.close()


@router.delete("/api/projects/{project_id}")
async def api_delete_project(project_id: int):
    """删除项目（保留关联数据）。"""
    db = SessionLocal()
    try:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            return {"error": "项目不存在"}
        db.delete(p)
        db.commit()
        return {"success": True}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


# ═══════════════════════════════════════════════════════
# 项目看板 API
# ═══════════════════════════════════════════════════════

@router.get("/api/projects/{project_id}/kanban")
async def api_project_kanban(project_id: int):
    """项目看板核心数据。"""
    db = SessionLocal()
    try:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            return {"error": "项目不存在"}

        # 词根
        roots = db.query(WordRoot).filter(
            WordRoot.batch_label == p.keyword_batch_label,
            WordRoot.domain == p.domain,
        ).order_by(WordRoot.frequency.desc()).all()

        # 月份列表
        kw_periods = db.query(SellerspriteKeyword.data_period).filter(
            SellerspriteKeyword.domain == p.domain,
            SellerspriteKeyword.batch_id == p.keyword_batch_label,
            SellerspriteKeyword.data_period != "",
        ).distinct().all()
        prod_periods = db.query(SellerspriteProduct.data_period).filter(
            SellerspriteProduct.domain == p.domain,
            SellerspriteProduct.batch_id == p.product_batch_label,
            SellerspriteProduct.data_period != "",
        ).distinct().all()
        all_periods = sorted(set(r[0] for r in kw_periods + prod_periods if r[0]))

        return {
            "project": {"id": p.id, "name": p.name, "domain": p.domain,
                        "keyword_batch_label": p.keyword_batch_label,
                        "product_batch_label": p.product_batch_label},
            "roots": [{"word": r.word, "frequency": r.frequency} for r in roots],
            "periods": all_periods,
        }
    finally:
        db.close()


@router.post("/api/projects/{project_id}/filter")
async def api_project_filter(project_id: int, body: dict):
    """按词根组合筛选关键词和 ASIN。"""
    db = SessionLocal()
    try:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            return {"error": "项目不存在"}

        domain = p.domain
        root_words = body.get("root_words", [])
        data_period = body.get("data_period", "")

        kw_query = db.query(SellerspriteKeyword).filter(
            SellerspriteKeyword.domain == domain,
            SellerspriteKeyword.batch_id == p.keyword_batch_label,
        )
        if data_period:
            kw_query = kw_query.filter(SellerspriteKeyword.data_period == data_period)
        kw_rows = kw_query.all()

        stop_words = _get_stop_words()
        matched_keywords = []
        for row in kw_rows:
            if root_words:
                kw_lower = row.keyword.lower()
                words = set(w for w in re.split(r"[^a-z0-9]+", kw_lower) if len(w) >= 2 and w not in stop_words)
                if not all(r in words for r in root_words):
                    continue
            matched_keywords.append(row)

        matched_asins_set = set()
        for row in matched_keywords:
            if row.top10_asins:
                for a in row.top10_asins.split(","):
                    a = a.strip()
                    if a.startswith("B0"):
                        matched_asins_set.add(a)

        products = {}
        if matched_asins_set:
            prod_query = db.query(SellerspriteProduct).filter(
                SellerspriteProduct.domain == domain,
                SellerspriteProduct.asin.in_(matched_asins_set),
            )
            if data_period:
                prod_query = prod_query.filter(SellerspriteProduct.data_period == data_period)
            products = {pr.asin: pr for pr in prod_query.all()}

        kw_data = []
        for row in matched_keywords:
            raw = json.loads(row.raw_response) if row.raw_response else {}
            kw_data.append({
                "keyword": row.keyword,
                "keyword_cn": row.keyword_cn or raw.get("keyword_cn", ""),
                "search_volume": row.search_volume or 0,
                "purchases": row.purchases_90d or 0,
                "clicks": row.clicks or 0,
                "purchase_rate": round((row.search_conversion_rate or 0), 1),
                "cpc": round(row.cpc_recommended or 0, 2),
                "avg_price": round(row.avg_price or 0, 2),
                "click_share": round((row.click_share or 0), 1),
                "conv_share": round((row.conv_share or 0), 1),
                "product_count": raw.get("product_count", 0),
                "review_count": raw.get("review_count", 0),
                "rating": raw.get("rating", 0),
            })

        asin_data = []
        for asin in sorted(matched_asins_set):
            pr = products.get(asin)
            raw = json.loads(pr.raw_response) if pr and pr.raw_response else {}
            asin_data.append({
                "asin": asin,
                "title": pr.title if pr else "",
                "brand": pr.brand if pr else "",
                "price": round(pr.price, 2) if pr and pr.price else 0,
                "monthly_sales": pr.monthly_sales if pr else 0,
                "monthly_revenue": round(pr.monthly_revenue, 2) if pr and pr.monthly_revenue else 0,
                "bsr": pr.main_bsr if pr else 0,
                "ratings_count": pr.ratings_count if pr else 0,
                "ratings": round(pr.ratings, 1) if pr and pr.ratings else 0,
                "online_days": pr.online_days if pr else 0,
                "is_fba": bool(pr.is_fba) if pr else False,
                "category": pr.main_category if pr else "",
                "seller_count": pr.seller_count if pr else 0,
            })

        summary = {
            "keyword_count": len(kw_data),
            "asin_count": len(asin_data),
            "total_sv": sum(k["search_volume"] for k in kw_data),
            "total_purchases": sum(k["purchases"] for k in kw_data),
            "avg_cpc": round(sum(k["cpc"] for k in kw_data) / max(len(kw_data), 1), 2),
            "avg_price": round(sum(k["avg_price"] for k in kw_data) / max(len(kw_data), 1), 2),
            "total_monthly_sales": sum(a["monthly_sales"] for a in asin_data),
            "total_monthly_revenue": round(sum(a["monthly_revenue"] for a in asin_data), 2),
        }

        return {"keywords": kw_data, "asins": asin_data, "summary": summary}
    finally:
        db.close()


@router.get("/api/projects/{project_id}/trends")
async def api_project_trends(project_id: int):
    """项目趋势数据：YoY 增长率 + 时序曲线。"""
    from sqlalchemy import func as sa_func

    db = SessionLocal()
    try:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            return {"error": "项目不存在"}

        # 按月聚合关键词数据
        rows = db.query(
            SellerspriteKeyword.data_period,
            sa_func.sum(SellerspriteKeyword.search_volume).label("total_sv"),
            sa_func.avg(SellerspriteKeyword.cpc_recommended).label("avg_cpc"),
            sa_func.avg(SellerspriteKeyword.purchases_90d).label("avg_purchases"),
            sa_func.avg(SellerspriteKeyword.clicks).label("avg_clicks"),
            sa_func.count(SellerspriteKeyword.id).label("kw_count"),
        ).filter(
            SellerspriteKeyword.batch_id == p.keyword_batch_label,
            SellerspriteKeyword.domain == p.domain,
            SellerspriteKeyword.data_period != "",
        ).group_by(SellerspriteKeyword.data_period).order_by(SellerspriteKeyword.data_period).all()

        # 按月聚合 Top10 产品销量
        # 先收集所有 ASIN
        all_kw = db.query(SellerspriteKeyword.top10_asins, SellerspriteKeyword.data_period).filter(
            SellerspriteKeyword.batch_id == p.keyword_batch_label,
            SellerspriteKeyword.domain == p.domain,
            SellerspriteKeyword.data_period != "",
            SellerspriteKeyword.top10_asins != "",
        ).all()
        period_asins = {}
        for top10, period in all_kw:
            if period not in period_asins:
                period_asins[period] = set()
            for a in top10.split(","):
                a = a.strip()
                if a.startswith("B0"):
                    period_asins[period].add(a)

        top10_sales_by_period = {}
        for period, asins in period_asins.items():
            sales_rows = db.query(
                SellerspriteProduct.monthly_sales
            ).filter(
                SellerspriteProduct.asin.in_(asins),
                SellerspriteProduct.domain == p.domain,
            ).order_by(SellerspriteProduct.monthly_sales.desc()).limit(10).all()
            top10_sales_by_period[period] = sum(s[0] or 0 for s in sales_rows)

        periods = []
        traffic_current = []
        traffic_yoy = []
        cpc_current = []
        cpc_yoy = []
        cpa_current = []
        cpa_yoy = []
        top10_current = []
        top10_yoy = []

        # 构建 year → period → value 映射用于 YoY 计算
        sv_by_period = {}
        cpc_by_period = {}
        cpa_by_period = {}
        for r in rows:
            pd_val = r.data_period
            sv_by_period[pd_val] = r.total_sv or 0
            cpc_by_period[pd_val] = r.avg_cpc or 0
            avg_p = r.avg_purchases or 0
            avg_c = r.avg_clicks or 0
            cpa_by_period[pd_val] = r.avg_cpc / (avg_p / avg_c) if avg_c and avg_p and r.avg_cpc else 0

        for r in rows:
            pd_val = r.data_period
            periods.append(pd_val)
            sv = r.total_sv or 0
            cpc = round(r.avg_cpc or 0, 2)
            cpa = round(cpa_by_period.get(pd_val, 0), 2)
            t10 = top10_sales_by_period.get(pd_val, 0)

            traffic_current.append(sv)
            cpc_current.append(cpc)
            cpa_current.append(cpa)
            top10_current.append(t10)

            # YoY: 找去年同月
            if len(pd_val) == 7 and "-" in pd_val:
                year, month = pd_val.split("-")
                prev_year = str(int(year) - 1)
                prev_period = f"{prev_year}-{month}"
            else:
                prev_period = None

            def _yoy(curr, prev):
                if prev and prev > 0:
                    return round((curr / prev - 1) * 100, 1)
                return None

            traffic_yoy.append(_yoy(sv, sv_by_period.get(prev_period)))
            cpc_yoy.append(_yoy(cpc, cpc_by_period.get(prev_period)))
            cpa_yoy.append(_yoy(cpa, cpa_by_period.get(prev_period)))
            t10_prev = top10_sales_by_period.get(prev_period)
            top10_yoy.append(_yoy(t10, t10_prev))

        return {
            "periods": periods,
            "traffic": {"current": traffic_current, "yoy_growth": traffic_yoy},
            "avg_cpc": {"current": cpc_current, "yoy_growth": cpc_yoy},
            "avg_cpa": {"current": cpa_current, "yoy_growth": cpa_yoy},
            "top10_sales": {"current": top10_current, "yoy_growth": top10_yoy},
        }
    finally:
        db.close()


# ═══════════════════════════════════════════════════════
# 项目赛道 CRUD
# ═══════════════════════════════════════════════════════

@router.post("/api/projects/{project_id}/niches")
async def api_project_save_niche(project_id: int, body: dict):
    """保存赛道到项目下。"""
    db = SessionLocal()
    try:
        nt = NicheTrack(
            name=body.get("name", ""),
            project_id=project_id,
            keyword_batch_label=body.get("keyword_batch_label", ""),
            product_batch_label=body.get("product_batch_label", ""),
            domain=body.get("domain", "US"),
            root_words=json.dumps(body.get("root_words", []), ensure_ascii=False),
            keyword_count=body.get("keyword_count", 0),
            asin_count=body.get("asin_count", 0),
            stats_snapshot=json.dumps(body.get("stats_snapshot", {}), ensure_ascii=False),
        )
        db.add(nt)
        db.commit()
        db.refresh(nt)
        return {"success": True, "id": nt.id}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


@router.get("/api/projects/{project_id}/niches")
async def api_project_list_niches(project_id: int):
    """列出项目下所有已保存赛道。"""
    db = SessionLocal()
    try:
        rows = db.query(NicheTrack).filter(
            NicheTrack.project_id == project_id,
        ).order_by(NicheTrack.created_at.desc()).all()
        return {
            "niches": [{
                "id": r.id,
                "name": r.name,
                "root_words": json.loads(r.root_words) if r.root_words else [],
                "keyword_count": r.keyword_count,
                "asin_count": r.asin_count,
                "stats_snapshot": json.loads(r.stats_snapshot) if r.stats_snapshot else {},
                "created_at": r.created_at.isoformat() if r.created_at else "",
            } for r in rows],
        }
    finally:
        db.close()


@router.get("/api/projects/{project_id}/niches/{niche_id}")
async def api_project_get_niche(project_id: int, niche_id: int):
    """获取赛道详情。"""
    db = SessionLocal()
    try:
        r = db.query(NicheTrack).filter(NicheTrack.id == niche_id, NicheTrack.project_id == project_id).first()
        if not r:
            return {"error": "赛道不存在"}
        return {
            "id": r.id, "name": r.name,
            "keyword_batch_label": r.keyword_batch_label,
            "product_batch_label": r.product_batch_label,
            "domain": r.domain,
            "root_words": json.loads(r.root_words) if r.root_words else [],
            "keyword_count": r.keyword_count,
            "asin_count": r.asin_count,
            "stats_snapshot": json.loads(r.stats_snapshot) if r.stats_snapshot else {},
            "created_at": r.created_at.isoformat() if r.created_at else "",
        }
    finally:
        db.close()


@router.delete("/api/projects/{project_id}/niches/{niche_id}")
async def api_project_delete_niche(project_id: int, niche_id: int):
    """删除赛道。"""
    db = SessionLocal()
    try:
        r = db.query(NicheTrack).filter(NicheTrack.id == niche_id, NicheTrack.project_id == project_id).first()
        if not r:
            return {"error": "赛道不存在"}
        db.delete(r)
        db.commit()
        return {"success": True}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}
    finally:
        db.close()


# ── 停用词表 ──────────────────────────────────────
def _get_stop_words() -> set:
    return {
        "a", "an", "the", "and", "or", "but", "in", "on", "at", "to", "for",
        "of", "with", "by", "from", "up", "about", "into", "through", "during",
        "before", "after", "above", "below", "between", "out", "off", "over",
        "under", "again", "further", "then", "once", "here", "there", "when",
        "where", "why", "how", "all", "both", "each", "every", "few", "more",
        "most", "other", "some", "such", "no", "nor", "not", "only", "own",
        "same", "so", "than", "too", "very", "just", "because", "as", "until",
        "while", "is", "am", "are", "was", "were", "be", "been", "being",
        "have", "has", "had", "do", "does", "did", "will", "would", "shall",
        "should", "may", "might", "must", "can", "could", "i", "me", "my",
        "we", "our", "you", "your", "he", "she", "it", "its", "they", "them",
        "their", "this", "that", "these", "those", "his", "her", "him",
    }

