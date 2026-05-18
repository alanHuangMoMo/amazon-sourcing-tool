"""
赛道递归分解引擎 — Recursive Market Decomposition Engine

═══════════════════
统一数据结构
═══════════════════
关键词元组: (keyword: str, sv: int, cpc: float, sales: int)
转化率 = sales / sv（sv>0时计算，否则0）
无论来源是 Sorftime 拓词还是文件上传，最终都归一化为这四个字段。

═══════════════════
总体逻辑
═══════════════════
输入任意关键词集合 → 词根共现图(Jaccard) → 贪心社区检测
→ LLM 判断每簇是"属性维度"或"子市场" → 子市场递归分解
→ 属性维度做指标汇总 → 输出市场结构树

两级深度：Level 0 LLM自由判断, Level 1+ 强制属性维度

═══════════════════
数据流
═══════════════════
来源A: 种子词 → Sorftime KeywordExtends → 字段映射 → 归一化
来源B: 上传文件 → LLM字段映射 → 归一化
归一化列表 → extract_roots() → cluster_cooccurrence()
→ llm_judge() → sub_market递归 / attribute汇总
"""
import sys, os as _os
sys.stdout.reconfigure(encoding='utf-8')

import asyncio, aiohttp, json, re, sqlite3, subprocess, tempfile
from collections import defaultdict
from itertools import combinations
from pathlib import Path
from datetime import datetime
from typing import Callable, Awaitable

# ═══ 可运行时注入的全局配置 ═══
API_KEY = _os.environ.get("DEEPSEEK_KEY", "sk-1fce6b2a9f7844d1938aa3ed512dbcde")
API_URL = "https://api.deepseek.com/v1/chat/completions"
MODEL = "deepseek-v4-flash"
LLM_TIMEOUT = 30
LLM_MAX_RETRIES = 1
DB_PATH = Path(__file__).parent.parent / "data" / "sourcing.db"
SORFTIME_PATH = "C:/Users/alanh/AppData/Roaming/npm/sorftime"
SORFTIME_DEFAULT_DOMAIN = 1
SORFTIME_PAGE_SIZE = 200

MAX_DEPTH = 2
MIN_KEYWORDS_FOR_DECOMPOSE = 10
MIN_KWS_PER_CLUSTER = 5
MIN_ROOT_COUNT = 2
MIN_JACCARD_SIM = 0.10
MAX_CLUSTERS_PER_LEVEL = 6

STOP_WORDS = {
    'a','an','the','is','are','was','were','be','been','being',
    'have','has','had','having','do','does','did','doing',
    'will','would','shall','should','can','could','may','might','must',
    'i','me','my','we','our','us','you','your','he','she','it','its',
    'they','them','their','this','that','these','those',
    'in','on','at','to','for','of','with','by','from',
    'and','or','but','not','no','nor','if','so','as','than',
    'also','too','very','just','about','up','out','down','off','over','under',
    'again','all','each','every','both','few','more','most',
    'other','some','such','only','own','same','new','now',
    'then','here','there','when','where','why','how','which','who','what','whom',
    'one','two','three','first','last','get','got','go','going',
    'into','onto','after','before','during','without','within',
    'per','like','much','many','any','been','still','well',
    'back','also','even','already','yet',
}

# Progress callback: async def(msg: str, pct: int)
ProgressFn = Callable[[str, int], Awaitable[None]]


# ════════════════ Phase 0: 字段映射 ════════════════

FIELD_MAP_PROMPT = """你是数据字段映射专家。给你一个文件的前几行表头和样本数据，
请将源字段映射到标准四字段：

标准字段：
- keyword: 关键词/搜索词/产品名
- sv: 月搜索量/搜索量
- cpc: 每次点击费用/平均CPC
- sales: 月销量/购买量

输出纯JSON（不要markdown）：
{"keyword": "源列名", "sv": "源列名或null", "cpc": "源列名或null", "sales": "源列名或null"}"""


async def map_fields_with_llm(headers: list[str], sample_rows: list[list[str]]) -> dict:
    """用 LLM 将上传文件的列映射到标准字段。"""
    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": FIELD_MAP_PROMPT},
            {"role": "user", "content": json.dumps({
                "headers": headers,
                "sample": sample_rows[:5],
            }, ensure_ascii=False)},
        ],
        "temperature": 0.0,
        "max_tokens": 300,
        "response_format": {"type": "json_object"},
    }
    headers_auth = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
    try:
        async with aiohttp.ClientSession(
            timeout=aiohttp.ClientTimeout(total=15)
        ) as session:
            async with session.post(API_URL, json=payload, headers=headers_auth) as resp:
                r = await resp.json()
                content = r["choices"][0]["message"]["content"].strip()
                if content.startswith("```"):
                    content = re.sub(r'^```\w*\n?', '', content)
                    content = re.sub(r'\n?```$', '', content)
                return json.loads(content)
    except Exception:
        return {"keyword": None, "sv": None, "cpc": None, "sales": None}


# ════════════════ Phase 0b: Sorftime 拓词 ════════════════

def sorftime_extend(seeds: list[str], domain: int = None, page_size: int = None,
                    progress: ProgressFn = None) -> list[dict]:
    """调用 Sorftime CLI KeywordExtends，返回原始 API 数据。"""
    if not seeds:
        raise ValueError("种子词列表为空")
    if domain is None:
        domain = SORFTIME_DEFAULT_DOMAIN
    if page_size is None:
        page_size = SORFTIME_PAGE_SIZE

    all_items = []
    for seed in seeds:
        params = json.dumps({"keyword": seed, "pageIndex": 1, "pageSize": page_size})
        outfile = tempfile.mktemp(suffix=".json")
        try:
            cmd = (
                f'bash {SORFTIME_PATH} api KeywordExtends '
                f'\'{params}\' --domain {domain} > "{outfile}" 2>&1'
            )
            subprocess.run(cmd, shell=True, timeout=45)
            if not Path(outfile).exists() or Path(outfile).stat().st_size == 0:
                print(f"  [WARN] '{seed}' 无输出")
                continue
            with open(outfile, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
            idx = text.find("{")
            if idx < 0:
                print(f"  [WARN] '{seed}' 响应中无JSON")
                continue
            data = json.loads(text[idx:])
            if data.get("Code") == 0:
                items = data.get("Data", [])
                all_items.extend(items)
                print(f"  '{seed}': {len(items)} 条, 剩余 {data.get('RequestLeft','?')} 次")
            else:
                print(f"  [WARN] '{seed}' Code={data.get('Code')}: {data.get('Message')}")
        except subprocess.TimeoutExpired:
            print(f"  [WARN] '{seed}' 超时")
        except json.JSONDecodeError as e:
            print(f"  [WARN] '{seed}' JSON解析失败: {e}")
        except Exception as e:
            print(f"  [WARN] '{seed}' {e}")
        finally:
            try:
                _os.unlink(outfile)
            except OSError:
                pass

    if not all_items:
        raise RuntimeError("所有种子词拓词失败。可能原因: 非ABA词 / 请求次数不足 / 网络问题")
    return all_items


def normalize_sorftime(items: list[dict], seeds: list[str]) -> list[tuple]:
    """
    将 Sorftime 返回的丰富字段 → 统一四字段。
    Sorftime 字段: Keyword, SearchVolume, Cpc, SalesVolumeOf90D
    """
    seed_set = {s.lower().strip() for s in seeds}
    result = []
    for it in items:
        kw = (it.get("Keyword") or "").strip()
        if not kw or kw.lower() in seed_set:
            continue
        sv = int(it.get("SearchVolume") or 0)
        cpc = float(it.get("Cpc") or 0)
        sales = int(it.get("SalesVolumeOf90D") or 0)
        result.append((kw, sv, cpc, sales))
    return result


# ════════════════ Phase 0c: 文件解析 ════════════════

def parse_uploaded_file(filepath: str) -> list[tuple]:
    """
    解析上传文件 → 归一化为 (kw, sv, cpc, sales) 列表。
    先用 LLM 识别字段映射，再解析数据行。
    """
    ext = Path(filepath).suffix.lower()
    if ext not in ('.csv', '.xlsx', '.xls'):
        raise RuntimeError(f"不支持的文件格式 '{ext}'。支持 CSV / Excel (.xlsx/.xls)")

    rows_raw = []
    if ext == '.csv':
        import csv as _csv
        with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = _csv.DictReader(f)
            for r in reader:
                rows_raw.append(r)
    else:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("缺少 openpyxl。运行: uv add openpyxl")
        wb = load_workbook(filepath, read_only=True)
        sheet = wb.active
        rows_iter = sheet.iter_rows()
        header_row = next(rows_iter, None)
        if not header_row:
            raise RuntimeError("文件为空")
        headers = [str(c.value or '').strip() for c in header_row]
        for row in rows_iter:
            vals = [str(c.value or '').strip() for c in row]
            rows_raw.append(dict(zip(headers, vals)))
        wb.close()

    if not rows_raw:
        return []

    headers = [str(h).strip() for h in rows_raw[0].keys() if h is not None and str(h).strip()]

    # 简单规则映射列名（中英文常见表头）
    mapping = {"keyword": None, "sv": None, "cpc": None, "sales": None}
    for h in headers:
        hl = h.lower().replace(' ', '').replace('_', '')
        if not mapping["keyword"] and any(x in hl for x in ['keyword', '关键词', '搜索词', 'searchterm', 'search_term', 'query']):
            mapping["keyword"] = h
        if not mapping["sv"] and any(x in hl for x in ['searchvolume', '搜索量', '月搜索量', 'sv', 'volume']):
            mapping["sv"] = h
        if not mapping["cpc"] and ('cpc' in hl or 'ppc' in hl or '竞价' in hl):
            mapping["cpc"] = h
        if not mapping["sales"] and any(x in hl for x in ['sales', '购买量', '月销量', '月购买量', 'salesvolume', '销量']):
            mapping["sales"] = h

    # Fallback
    if not mapping["keyword"]:
        mapping["keyword"] = headers[0]

    def _num(v):
        if v is None:
            return 0
        try:
            return float(str(v).replace('$', '').replace('%', '').replace(',', '').strip())
        except (ValueError, TypeError):
            return 0

    result = []
    for r in rows_raw:
        kw = str(r.get(mapping["keyword"], '')).strip()
        if not kw or len(kw) < 2:
            continue
        sv = int(_num(r.get(mapping["sv"], 0))) if mapping["sv"] else 0
        cpc = _num(r.get(mapping["cpc"], 0)) if mapping["cpc"] else 0
        sales = int(_num(r.get(mapping["sales"], 0))) if mapping["sales"] else 0
        result.append((kw, sv, cpc, sales))
    return result


# ════════════════ Phase 1: 词根提取 ════════════════

def extract_roots(keywords: list[tuple]) -> tuple:
    """返回 (kw_roots, root_docs)"""
    root_docs = defaultdict(list)
    kw_roots = []
    for idx, row in enumerate(keywords):
        kw = str(row[0]).lower() if row[0] else ""
        words = set()
        for w in kw.split():
            clean = re.sub(r'[^a-z]', '', w)
            if len(clean) >= 2 and clean not in STOP_WORDS:
                words.add(clean)
                root_docs[clean].append(idx)
        kw_roots.append(words)
    return kw_roots, root_docs


# ════════════════ Phase 1b: 共现图社区检测 ════════════════

def cluster_cooccurrence(kw_roots: list, root_docs: dict,
                         min_count: int = MIN_ROOT_COUNT,
                         min_sim: float = MIN_JACCARD_SIM) -> list:
    """词根共现图 → Jaccard边权 → 贪心合并 → 词根簇列表"""
    valid = {r for r, idxs in root_docs.items() if len(idxs) >= min_count}
    valid_roots = sorted(valid)
    if not valid_roots:
        return []
    root_idx = {r: i for i, r in enumerate(valid_roots)}
    n = len(valid_roots)
    for i in range(len(kw_roots)):
        kw_roots[i] = kw_roots[i] & valid

    doc_counts = {r: len(root_docs[r]) for r in valid_roots}
    edges = defaultdict(float)
    for roots in kw_roots:
        vk = list(roots)
        if len(vk) < 2:
            continue
        for a, b in combinations(sorted(vk), 2):
            if a in root_idx and b in root_idx:
                i, j = root_idx[a], root_idx[b]
                edges[(min(i, j), max(i, j))] += 1

    for (i, j), cooc in edges.items():
        ra, rb = valid_roots[i], valid_roots[j]
        denom = doc_counts[ra] + doc_counts[rb] - cooc
        edges[(i, j)] = cooc / denom if denom > 0 else 0

    parent = list(range(n))
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(x, y):
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py
            return True
        return False

    for (i, j), w in sorted(edges.items(), key=lambda x: -x[1]):
        if w < min_sim:
            break
        union(i, j)

    comms = defaultdict(list)
    for i in range(n):
        comms[find(i)].append(i)

    results = []
    for leader, members in comms.items():
        if len(members) < 2:
            continue
        comm_roots = [valid_roots[i] for i in members]
        results.append({
            "id": f"comm_{leader}",
            "roots": comm_roots,
            "n_roots": len(comm_roots),
            "top_roots": sorted(comm_roots,
                key=lambda r: len(root_docs.get(r, [])), reverse=True)[:12]
        })
    results.sort(key=lambda x: -x["n_roots"])
    return results


def get_community_keywords(comm_roots: list, all_keywords: list[tuple]) -> list[tuple]:
    """获取含该社区任意词根的关键词子集"""
    root_set = set(comm_roots)
    matched = []
    for row in all_keywords:
        kw = str(row[0]).lower() if row[0] else ""
        words = set(re.sub(r'[^a-z]', ' ', kw).split())
        if words & root_set:
            matched.append(row)
    return matched


# ════════════════ Phase 2: LLM 簇判断 ════════════════

JUDGE_PROMPT = """你是电商关键词分析专家。判断词根簇类型。

sub_market: 定义了产品子市场（如mug/cup/tumbler→杯具，shirt/tee→服装）
attribute_dimension: 同一属性不同取值（如black/white/red→颜色，cotton/polyester→材质）
leaf: 太小或无法判断

输出纯JSON:
{"type":"sub_market|attribute_dimension|leaf","name":"中文名","name_en":"en","roots_for_next_level":[],"attribute_values":[],"attribute_name":"","core_terms":[],"non_english":[],"note":""}"""


async def _llm_api(payload: dict, retries: int = LLM_MAX_RETRIES) -> dict:
    headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
    last_err = None
    for attempt in range(retries + 1):
        try:
            t = aiohttp.ClientTimeout(total=LLM_TIMEOUT)
            async with aiohttp.ClientSession(timeout=t) as s:
                async with s.post(API_URL, json=payload, headers=headers) as r:
                    if r.status != 200:
                        raise RuntimeError(f"HTTP {r.status}")
                    d = await r.json()
                    if "error" in d:
                        raise RuntimeError(f"API: {d['error']}")
                    c = d["choices"][0]["message"]["content"].strip()
                    if c.startswith("```"):
                        c = re.sub(r'^```\w*\n?', '', c)
                        c = re.sub(r'\n?```$', '', c)
                    return json.loads(c)
        except (aiohttp.ClientError, asyncio.TimeoutError, json.JSONDecodeError) as e:
            last_err = e
            if attempt < retries:
                await asyncio.sleep(1 * (attempt + 1))
    raise RuntimeError(f"LLM调用失败(重试{retries}次): {last_err}")


async def llm_judge(community: dict, kw_sample: list[tuple]) -> dict:
    roots = community.get("top_roots", community.get("roots", []))
    samples = [str(r[0]) for r in kw_sample[:25] if r[0]]
    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": JUDGE_PROMPT},
            {"role": "user", "content": json.dumps(
                {"roots": roots, "keyword_samples": samples}, ensure_ascii=False
            )},
        ],
        "temperature": 0.1, "max_tokens": 1000,
        "response_format": {"type": "json_object"},
    }
    try:
        return await _llm_api(payload)
    except Exception as e:
        print(f"  [LLM fallback] {e}")
        if len(roots) >= 3:
            return {"type": "sub_market", "name": f"簇({len(roots)}词)",
                    "core_terms": roots[:5], "note": f"降级: {e}"}
        return {"type": "leaf", "name": f"小簇({len(roots)}词)", "note": f"降级: {e}"}


# ════════════════ Phase 3: 指标汇总 ════════════════

def aggregate_by_root(kw_list: list[tuple], root: str) -> dict | None:
    """
    对包含指定词根的关键词子集汇总指标。
    统一结构: sv, cpc, sales → 派生 conv = sales/sv
    需求强度 = sv / avg_products（sv越高越蓝海，乘积越小越红海）
    """
    if not kw_list or not root:
        return None
    matched = [r for r in kw_list
               if root in set(re.sub(r'[^a-z]', ' ', str(r[0]).lower()).split())]
    if len(matched) == 0:
        return None

    total_sv = sum(r[1] or 0 for r in matched)
    total_sales = sum(r[3] or 0 for r in matched)
    cpcs = [r[2] for r in matched if r[2] and r[2] > 0]
    avg_cpc = round(sum(cpcs) / len(cpcs), 2) if cpcs else 0
    conv = round(total_sales / total_sv, 4) if total_sv > 0 else 0

    return {
        "value": root,
        "n_keywords": len(matched),
        "total_sv": total_sv,
        "total_sales": total_sales,
        "avg_cpc": avg_cpc,
        "conv": conv,          # 转化率 = sales/sv
        "intensity": round(total_sv / max(len(matched), 1), 1),
    }


# ════════════════ Phase 4: 递归分解 ════════════════

async def decompose(keywords: list[tuple], depth: int = 0, parent_name: str = "root",
                    max_depth: int = MAX_DEPTH,
                    progress: ProgressFn = None) -> dict:
    indent = "  " * depth
    n_kw = len(keywords)
    print(f"{indent}[{parent_name}] depth={depth}, keywords={n_kw}")

    if progress:
        await progress(f"分析 {parent_name} ({n_kw}词)", min(20 + depth * 30, 90))

    # 停止条件
    if n_kw < MIN_KEYWORDS_FOR_DECOMPOSE:
        return {"type": "leaf", "name": parent_name, "n_keywords": n_kw}
    if depth >= max_depth:
        return {"type": "leaf", "name": parent_name, "n_keywords": n_kw}

    try:
        kw_roots, root_docs = extract_roots(keywords)
    except Exception as e:
        return {"type": "leaf", "name": parent_name,
                "n_keywords": n_kw, "note": f"词根提取失败: {e}"}

    communities = cluster_cooccurrence(kw_roots, root_docs)
    if not communities:
        return {"type": "leaf", "name": parent_name,
                "n_keywords": n_kw, "note": "未发现社区结构"}

    print(f"{indent}  {len(communities)} communities")

    children = []
    for comm in communities[:MAX_CLUSTERS_PER_LEVEL]:
        comm_kws = get_community_keywords(comm["roots"], keywords)
        if len(comm_kws) < MIN_KWS_PER_CLUSTER:
            continue

        try:
            judgment = await llm_judge(comm, comm_kws)
        except Exception as e:
            judgment = {"type": "leaf", "name": "错误簇", "note": str(e)}

        jtype = judgment.get("type", "leaf")
        jname = judgment.get("name", f"cluster_{comm['id']}")
        core = set(judgment.get("core_terms", []))
        print(f"{indent}  -> {jname} [{jtype}] ({len(comm_kws)}kws)")

        # 深度>=1 强制 attribute
        if depth >= 1 and jtype == "sub_market":
            jtype = "attribute_dimension"
            proots = [r for r in comm.get("top_roots", []) if r in core][:10]
            if not proots:
                proots = comm.get("top_roots", [])[:8]
            if not judgment.get("attribute_name"):
                judgment["attribute_name"] = jname
            if not judgment.get("attribute_values"):
                judgment["attribute_values"] = proots

        if jtype == "attribute_dimension":
            vals = []
            for v in judgment.get("attribute_values", comm.get("top_roots", [])[:10]):
                agg = aggregate_by_root(comm_kws, v)
                if agg and agg["n_keywords"] >= 2:
                    vals.append(agg)
            vals.sort(key=lambda x: -x["total_sv"])
            children.append({"type": "attribute",
                "name": judgment.get("attribute_name", jname),
                "name_en": judgment.get("name_en", ""),
                "values": vals, "note": judgment.get("note", "")})

        elif jtype == "sub_market":
            sub_roots = set(judgment.get("roots_for_next_level", comm["roots"]))
            sub_kws = [r for r in comm_kws
                       if set(re.sub(r'[^a-z]', ' ', str(r[0]).lower()).split()) & sub_roots]
            if len(sub_kws) >= MIN_KEYWORDS_FOR_DECOMPOSE:
                try:
                    sub = await decompose(sub_kws, depth + 1, jname, max_depth, progress)
                except Exception as e:
                    sub = {"type": "leaf", "name": jname,
                           "n_keywords": len(sub_kws), "note": str(e)}
                sub_ch = sub.get("children", [sub]) if sub.get("type") != "leaf" else []
                children.append({"type": "market", "name": jname,
                    "name_en": judgment.get("name_en", ""),
                    "n_keywords": len(sub_kws),
                    "total_sv": sum(r[1] or 0 for r in sub_kws),
                    "children": sub_ch,
                    "core_terms": judgment.get("core_terms", []),
                    "note": judgment.get("note", "")})
            else:
                children.append({"type": "leaf", "name": jname,
                    "n_keywords": len(sub_kws),
                    "note": f"关键词不足({len(sub_kws)})"})
        else:
            children.append({"type": "leaf", "name": jname,
                "n_keywords": len(comm_kws), "note": judgment.get("note", "")})

    if not children:
        children.append({"type": "leaf", "name": parent_name,
            "n_keywords": n_kw, "note": "无有效簇"})

    return {"type": "market", "name": parent_name,
        "n_keywords": n_kw,
        "total_sv": sum(r[1] or 0 for r in keywords),
        "children": children}


# ════════════════ 顶层入口 ════════════════

async def run_exploration(seeds: list[str], domain: int = 1,
                          progress: ProgressFn = None) -> dict:
    """种子词模式"""
    if not seeds:
        raise ValueError("种子词为空")
    if len(seeds) > 5:
        raise ValueError(f"最多5个种子词，当前{len(seeds)}个")

    if progress:
        await progress(f"Sorftime 拓词中 ({', '.join(seeds)})...", 5)

    print(f"\n=== Auto Explorer: {seeds} ===\n")
    items = sorftime_extend(seeds, domain, progress=progress)
    print(f"  Total {len(items)} items from Sorftime")

    if progress:
        await progress("数据归一化...", 15)

    keywords = normalize_sorftime(items, seeds)
    print(f"  Normalized {len(keywords)} keywords")

    if len(keywords) < MIN_KEYWORDS_FOR_DECOMPOSE:
        return {"error": f"有效关键词不足({len(keywords)})",
                "total_keywords": len(keywords), "seeds": seeds}

    if progress:
        await progress(f"开始递归分解 ({len(keywords)}词)...", 20)

    tree = await decompose(keywords, 0, "all", MAX_DEPTH, progress)
    tree["seeds"] = seeds
    tree["total_keywords"] = len(keywords)
    tree["total_sv"] = sum(r[1] or 0 for r in keywords)
    tree["timestamp"] = datetime.now().isoformat()

    if progress:
        await progress("完成", 100)

    print(f"\n=== Done: {tree.get('total_keywords')} keywords -> tree ===\n")
    return tree


async def run_exploration_from_file(filepath: str, source_name: str = "upload",
                                    progress: ProgressFn = None) -> dict:
    """文件上传模式"""
    if not filepath or not Path(filepath).exists():
        raise FileNotFoundError(f"文件不存在: {filepath}")

    if progress:
        await progress("解析文件...", 5)

    print(f"\n=== Auto Explorer (file): {source_name} ===\n")
    keywords = parse_uploaded_file(filepath)
    print(f"  Parsed {len(keywords)} keywords")

    if len(keywords) < MIN_KEYWORDS_FOR_DECOMPOSE:
        return {"error": f"文件仅有{len(keywords)}条有效关键词（需>={MIN_KEYWORDS_FOR_DECOMPOSE}）",
                "total_keywords": len(keywords)}

    if progress:
        await progress(f"开始递归分解 ({len(keywords)}词)...", 20)

    tree = await decompose(keywords, 0, "all", MAX_DEPTH, progress)
    tree["seeds"] = [source_name]
    tree["total_keywords"] = len(keywords)
    tree["total_sv"] = sum(r[1] or 0 for r in keywords)
    tree["timestamp"] = datetime.now().isoformat()

    if progress:
        await progress("完成", 100)

    print(f"\n=== Done: {tree.get('total_keywords')} keywords -> tree ===\n")
    return tree
