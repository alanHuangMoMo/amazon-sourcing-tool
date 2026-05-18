"""
Phase 2: 关键词词库导出 → 前10 ASIN 回填子赛道。

流程:
  1. 从 final_sub_niche_kw 读取所有关键词（按子赛道分组）
  2. 全局去重后分批（≤2000 词/批）
  3. 调 keyword_store_batch.py 导出
  4. 解析 Excel → 提取每关键词的 top10_asins
  5. 通过关键词→子赛道映射回填 ASIN

用法:
  python batch_kw_to_asin.py
  python batch_kw_to_asin.py --resume
  python batch_kw_to_asin.py --dry-run
"""

import argparse
import json
import os
import sqlite3
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
STORE_SCRIPT = PROJECT_ROOT / "sellersprite_keyword_store" / "keyword_store_batch.py"
SCRIPT_PYTHON = r"C:\Users\alanh\AppData\Local\Programs\Python\Python313\python.exe"
DATA_DIR = Path(__file__).resolve().parent / "data"
BATCH_DIR = DATA_DIR / "kw_batches"
OUTPUT_DIR = DATA_DIR / "kw_outputs"

BATCH_MAX = 2000
MARKET = "CA"
EXPORT_QUOTA = 50
EXPORT_KEY = "keyword_store"
EXPORT_COUNT_FILE = DATA_DIR / "sellersprite_export_count.json"


def load_export_count() -> dict:
    today = time.strftime("%Y-%m-%d")
    if EXPORT_COUNT_FILE.exists():
        data = json.loads(EXPORT_COUNT_FILE.read_text())
        if data.get("date") != today:
            return {"date": today, "comparison": 0, "keyword_store": 0}
        data.setdefault(EXPORT_KEY, 0)
        return data
    return {"date": today, "comparison": 0, "keyword_store": 0}


def increment_export_count():
    data = load_export_count()
    data[EXPORT_KEY] = data.get(EXPORT_KEY, 0) + 1
    EXPORT_COUNT_FILE.parent.mkdir(parents=True, exist_ok=True)
    EXPORT_COUNT_FILE.write_text(json.dumps(data))


def check_quota() -> bool:
    data = load_export_count()
    cnt = data.get(EXPORT_KEY, 0)
    remain = EXPORT_QUOTA - cnt
    print(f"今日 {EXPORT_KEY}: {cnt}/{EXPORT_QUOTA} (剩余 {remain})")
    if remain <= 0:
        print(f"!!! 今日额度用完，请明天继续")
        return False
    return True


def load_db():
    db_path = DATA_DIR / "sourcing.db"
    db = sqlite3.connect(str(db_path))
    db.execute("PRAGMA journal_mode=WAL")
    return db


def get_keyword_batches(db) -> list[list[str]]:
    """读取所有唯一关键词，分成 ≤2000 的批次"""
    rows = db.execute("SELECT DISTINCT keyword FROM final_sub_niche_kw ORDER BY keyword").fetchall()
    all_kws = [r[0] for r in rows]
    batches = []
    for i in range(0, len(all_kws), BATCH_MAX):
        batches.append(all_kws[i:i + BATCH_MAX])
    return batches


def build_kw_niche_map(db) -> dict[str, set[int]]:
    """关键词 → 子赛道 ID 集合"""
    rows = db.execute("SELECT keyword, sub_niche_id FROM final_sub_niche_kw").fetchall()
    m = defaultdict(set)
    for kw, sid in rows:
        m[kw].add(sid)
    return dict(m)


def run_keyword_store(keywords: list[str], batch_idx: int) -> Path | None:
    """调 keyword_store_batch.py，返回下载的 xlsx 路径"""
    batch_file = BATCH_DIR / f"kw_batch_{batch_idx:04d}.txt"
    batch_file.parent.mkdir(parents=True, exist_ok=True)
    batch_file.write_text("\n".join(keywords), encoding="utf-8")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    cmd = [
        SCRIPT_PYTHON, str(STORE_SCRIPT),
        str(batch_file),
        "--market", MARKET,
        "--prefix", f"KW-{batch_idx:04d}",
    ]

    print(f"\n{'='*60}")
    print(f"词库批次 {batch_idx}: {len(keywords)} 关键词")
    print(f"{'='*60}")

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=600,
            encoding="utf-8",
            errors="replace",
            cwd=str(PROJECT_ROOT / "sellersprite_keyword_store"),
            env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        stdout = result.stdout
        stderr = result.stderr
        if stderr:
            print(f"  [stderr]: {stderr[:500]}")

        lines = stdout.splitlines()
        for line in lines[-30:]:
            print(line)

        if result.returncode != 0:
            print(f"  词库批次 {batch_idx} 失败 (exit {result.returncode})")
            return None

        for line in lines:
            if "已保存:" in line or "文件:" in line:
                path_str = line.split(":", 1)[-1].strip()
                p = Path(path_str)
                if p.exists() and p.suffix == ".xlsx":
                    return p

        xlsx_files = sorted(OUTPUT_DIR.glob("*.xlsx"),
                           key=lambda f: f.stat().st_mtime, reverse=True)
        if xlsx_files:
            return xlsx_files[0]

        print(f"  词库批次 {batch_idx}: 找不到输出文件")
        return None

    except subprocess.TimeoutExpired:
        print(f"  词库批次 {batch_idx} 超时")
        return None


def parse_keyword_store(xlsx_path: Path) -> dict[str, list[str]]:
    """解析关键词词库导出 Excel，提取 {keyword: [asin, ...]} 从前10ASIN列"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 找列头
    kw_col = None
    asin_col = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if not v:
            continue
        h = str(v).strip()
        if h == "关键词":
            kw_col = col
        elif "前10" in h and "ASIN" in h:
            asin_col = col

    if not kw_col:
        print("  找不到「关键词」列")
        return {}
    if not asin_col:
        print("  找不到「前10ASIN」列")
        return {}

    result = {}
    for row in range(2, ws.max_row + 1):
        kw = ws.cell(row, kw_col).value
        asins_str = ws.cell(row, asin_col).value
        if not kw or not asins_str:
            continue
        kw = str(kw).strip()
        # 过滤系统提示行（如"以下为未获取到关键词数据的词..."）
        if len(kw) > 100 or "Unique Words" in kw or "词频统计" in kw:
            continue
        asins = [a.strip() for a in str(asins_str).split(",") if a.strip().startswith("B0")]
        if asins:
            result[kw] = asins

    return result


def store_asins(db, kw_asin_map: dict[str, list[str]],
                kw_niche_map: dict[str, set[int]]) -> tuple[int, int]:
    """回填 ASIN 到 final_sub_niche_asin"""
    inserted = 0
    skipped = 0
    for kw, asins in kw_asin_map.items():
        niche_ids = kw_niche_map.get(kw, set())
        if not niche_ids:
            continue
        for asin in asins:
            for sid in niche_ids:
                try:
                    db.execute(
                        "INSERT INTO final_sub_niche_asin (sub_niche_id, asin) VALUES (?, ?)",
                        (sid, asin),
                    )
                    inserted += 1
                except sqlite3.IntegrityError:
                    skipped += 1
    db.commit()
    return inserted, skipped


def save_progress(batch_idx: int, total: int):
    """记录完成进度。只允许递增，防止并发 session 互相覆盖"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    progress_file = DATA_DIR / "kw_expand_progress.json"
    data = load_progress()
    if data and data.get("last_completed_batch", -1) >= batch_idx:
        return
    progress_file.write_text(json.dumps({
        "last_completed_batch": batch_idx,
        "total_batches": total,
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }))


def load_progress():
    progress_file = DATA_DIR / "kw_expand_progress.json"
    if progress_file.exists():
        return json.loads(progress_file.read_text())
    return None


def _cleanup_chrome():
    try:
        subprocess.run(["taskkill", "/F", "/IM", "chrome.exe"],
                       capture_output=True, timeout=30)
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--batches", type=int, default=0, help="本次只跑 N 批，完成后自动停（0=不限）")
    args = parser.parse_args()

    db = load_db()

    batches = get_keyword_batches(db)
    print(f"共 {len(batches)} 词库批次, {sum(len(b) for b in batches)} 唯一关键词")

    kw_niche_map = build_kw_niche_map(db)
    print(f"关键词→子赛道映射: {len(kw_niche_map)} 词")

    if args.dry_run:
        for i, batch in enumerate(batches[:3]):
            # 统计这批词覆盖多少子赛道
            niches = set()
            for kw in batch:
                niches.update(kw_niche_map.get(kw, set()))
            print(f"  批次 {i}: {len(batch)} 词, 覆盖 {len(niches)} 子赛道")
        return

    start_batch = 0
    if args.resume:
        prog = load_progress()
        if prog:
            start_batch = prog["last_completed_batch"] + 1
            print(f"续跑: 从批次 {start_batch} 开始")

    total_inserted = 0
    total_skipped = 0
    failed_batches = []
    quota_exhausted = False

    for i in range(start_batch, len(batches)):
        batch = batches[i]

        if args.batches and (i - start_batch) >= args.batches:
            print(f"本次已跑 {args.batches} 批，交棒给下个 session")
            break

        if not check_quota():
            quota_exhausted = True
            break

        xlsx_path = run_keyword_store(batch, i)
        increment_export_count()

        if not xlsx_path:
            print(f"!!! 词库批次 {i} 失败，跳过继续")
            failed_batches.append(i)
            _cleanup_chrome()
            continue

        kw_asin_map = parse_keyword_store(xlsx_path)
        total_asins = sum(len(v) for v in kw_asin_map.values())
        print(f"  解析: {len(kw_asin_map)} 关键词, {total_asins} ASINs")

        inserted, skipped = store_asins(db, kw_asin_map, kw_niche_map)
        total_inserted += inserted
        total_skipped += skipped
        print(f"  入库: +{inserted} 新ASIN, {skipped} 已存在")

        save_progress(i, len(batches))
        _cleanup_chrome()

    print(f"\n{'='*60}")
    if quota_exhausted:
        data = load_export_count()
        print(f"今日额度用完 ({data[EXPORT_KEY]}/{EXPORT_QUOTA})，已暂停")
        print(f"当前进度: {load_progress()['last_completed_batch']+1}/{len(batches)} 批")
        print(f"明天 cron 自动续跑，或手动: python batch_kw_to_asin.py --resume")
    else:
        print(f"Phase 2 完成! 新增 ASIN: {total_inserted}, 跳过: {total_skipped}")
    if failed_batches:
        print(f"失败批次: {failed_batches}")

    new_asin_cnt = db.execute("SELECT COUNT(DISTINCT asin) FROM final_sub_niche_asin").fetchone()[0]
    print(f"final_sub_niche_asin 总计: {new_asin_cnt} 唯一 ASIN")
    db.close()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback, time
        tb = traceback.format_exc()
        print(f"FATAL: {tb[:1000]}")
        err_log = Path(__file__).resolve().parent / "data" / "kw_expand_errors.log"
        with open(err_log, "a", encoding="utf-8") as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"FATAL 崩溃 @ {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*60}\n")
            f.write(tb)
            f.write("\n")
        sys.exit(1)
