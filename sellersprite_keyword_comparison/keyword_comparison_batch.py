"""
卖家精灵流量词对比 → ASIN P1 流量入口
批量查询最多 10 个竞品 ASIN，提取每个 ASIN 的第1页关键词。

流程：登录 → 输 ASIN → 查询 → getByRole 点畅销变体 → 等导出按钮启用 → 导出 → 下载 → 清洗

用法:
  python keyword_comparison_batch.py asins.txt
  python keyword_comparison_batch.py asins.txt --market UK
  python keyword_comparison_batch.py asins.txt --own-asin B0XXX
"""

import argparse
import asyncio
import json
import random
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

# 确保项目根在 path，可以 import sellersprite_utils
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from playwright.async_api import async_playwright

# ---- 常量 ----
USER_DATA = Path.home() / ".sellersprite-comparison-profile"
COMPARISON_URL = "https://www.sellersprite.com/v3/keyword-comparison"
EXPORT_LOG_URL = "https://www.sellersprite.com/v2/export-log"
MAX_ASINS = 10

MARKET_MAP = {
    "US": "美国站", "CA": "加拿大", "JP": "日本站", "UK": "英国站",
    "DE": "德国站", "FR": "法国站", "IT": "意大利", "ES": "西班牙", "IN": "印度站",
}

DIMENSION_MAP = {
    "traffic_share": 1,    # 流量占比（默认）
    "weekly_exposure": 2,  # 周曝光量
    "natural_rank": 3,     # 自然排名
    "ad_rank": 4,          # 广告排名
    "conversion": 5,       # 转化效果
    "exposure_position": 6, # 曝光位置
}

LOGIN_URL_RE = re.compile(r'/(login|passport)')
CREDENTIALS_FILE = Path.home() / ".sellersprite-credentials"
PAGE1_RANK_LABEL = "主要流量词"


def load_credentials():
    try:
        data = json.loads(CREDENTIALS_FILE.read_text(encoding="utf-8"))
        return data["email"], data["password"]
    except Exception as e:
        print(f"无法读取凭据 {CREDENTIALS_FILE}: {e}")
        sys.exit(1)


async def human_delay(min_s=0.3, max_s=1.2):
    await asyncio.sleep(random.uniform(min_s, max_s))


async def is_on_login_page(page):
    return bool(LOGIN_URL_RE.search(page.url))


# ============================================================
#  登录
# ============================================================

async def auto_login(page):
    email, password = load_credentials()
    print("自动登录中...")
    await human_delay(1.0, 2.0)
    try:
        await page.wait_for_selector('input[name="email"], input[type="password"]', timeout=10000)
    except Exception:
        pass
    await human_delay(0.5, 1.0)
    try:
        await page.locator('input[name="email"]:visible').fill(email)
    except Exception:
        try:
            await page.locator('input[placeholder*="手机号"]:visible, input[placeholder*="邮箱"]:visible').fill(email)
        except Exception:
            pass
    try:
        await page.locator('input[type="password"]:visible').fill(password)
    except Exception:
        pass
    await human_delay(0.5, 1.0)
    try:
        await page.locator('button:visible').filter(has_text='登录').click(timeout=5000)
    except Exception:
        try:
            await page.locator('button[type="submit"]:visible').click(timeout=5000)
        except Exception:
            pass

    # 检测账密错误 / 验证码
    await asyncio.sleep(2)
    from sellersprite_utils.credential_check import detect_login_error
    err = await detect_login_error(page)
    if err:
        if "验证码" in err:
            print(f"\n>>> 检测到验证码，请手动完成滑块/图形验证 <<<")
            print("等待手动登录（最长 5 分钟）", end="", flush=True)
            for _ in range(300):
                await asyncio.sleep(1)
                if not await is_on_login_page(page):
                    print(" 登录成功")
                    return True
                print(".", end="", flush=True)
            print(" 超时")
            return False
        else:
            print(f"\n!!! 账密验证失败: {err}")
            print("停止执行，请更新账密后重试")
            sys.exit(1)

    print("等待登录完成", end="", flush=True)
    for _ in range(100):
        await asyncio.sleep(1)
        if not await is_on_login_page(page):
            print(" 成功")
            return True
        print(".", end="", flush=True)
    print(" 超时")
    return False


async def ensure_logged_in(page):
    await page.goto(COMPARISON_URL, timeout=30000)
    await page.wait_for_load_state("domcontentloaded")
    try:
        await page.wait_for_selector('button:has-text("立即查询")', timeout=10000)
    except Exception:
        pass
    await human_delay(1.0, 1.5)

    if await is_on_login_page(page):
        print("需要登录...")
        await page.goto("https://www.sellersprite.com/w/user/login", timeout=30000)
        await page.wait_for_load_state("domcontentloaded")
        if not await auto_login(page):
            sys.exit(1)
        await page.goto(COMPARISON_URL, timeout=30000)
        await page.wait_for_load_state("domcontentloaded")
        await human_delay(1.0, 1.5)
    else:
        is_guest = await page.evaluate(
            """() => {
                const t = document.body.innerText;
                return t.includes('未登录') || t.includes('游客');
            }"""
        )
        if is_guest:
            print("检测到游客，登录中...")
            await page.goto("https://www.sellersprite.com/w/user/login", timeout=30000)
            await page.wait_for_load_state("domcontentloaded")
            if not await auto_login(page):
                sys.exit(1)
            await page.goto(COMPARISON_URL, timeout=30000)
            await page.wait_for_load_state("domcontentloaded")
            await human_delay(1.0, 1.5)
        else:
            print("已登录")


# ============================================================
#  查询 + 变体弹窗 + 导出
# ============================================================

async def run_comparison(page, asins: list[str], own_asin: str = "", dimension: str = "traffic_share"):
    """输入 ASIN → 查询 → 处理变体弹窗 → 切维度 → 等导出按钮可用"""
    asin_str = ",".join(asins)
    print(f"  查询 {len(asins)} 个 ASIN: {asin_str[:60]}...")

    try:
        await page.locator('input[placeholder*="竞品ASIN"]').first.fill(asin_str)
    except Exception:
        pass
    if own_asin:
        try:
            await page.locator('input[placeholder*="自己的ASIN"]').first.fill(own_asin)
        except Exception:
            pass

    await human_delay(0.3, 0.8)

    # 点「立即查询」
    try:
        await page.get_by_role("button", name="立即查询").click(timeout=10000)
    except Exception:
        pass

    # 等变体弹窗出现 + 点「用畅销变体拓词」
    print("  检查变体弹窗...")
    variant_clicked = False
    for _ in range(30):
        # 用 evaluate 检查 + 点击（绕过 Playwright locator 问题）
        info = await page.evaluate(
            """() => {
                const btns = document.querySelectorAll('button');
                for (const b of btns) {
                    if (b.offsetWidth > 0 && b.innerText.includes('用畅销变体')) {
                        // 移除可能拦截点击的 overlay
                        const overlays = document.querySelectorAll('.v-modal, .el-overlay');
                        overlays.forEach(o => { if (o.style) o.style.display = 'none'; });
                        // 直接触发 Vue 组件方法
                        const parent = b.closest('.el-dialog, [role="dialog"]');
                        b.click();
                        return {clicked: true, text: b.innerText.trim().substring(0, 30)};
                    }
                }
                return {clicked: false};
            }"""
        )
        if info.get("clicked"):
            print(f"  已点击: {info.get('text')}")
            variant_clicked = True
            break
        # 超时前检查是否结果已加载
        rows = await page.evaluate("() => document.querySelectorAll('table tbody tr').length")
        if rows > 5:
            print("  结果已直接加载")
            break
        await asyncio.sleep(1)

    if not variant_clicked:
        print("  无变体弹窗")
    await asyncio.sleep(3)

    # 确保变体弹窗彻底消失（可能重新弹出）
    for _ in range(30):
        still_open = await page.evaluate(
            """() => {
                const btns = document.querySelectorAll('button');
                for (const b of btns) {
                    if (b.offsetWidth > 0 && b.innerText.includes('用畅销变体')) {
                        const overlays = document.querySelectorAll('.v-modal, .el-overlay');
                        overlays.forEach(o => { if (o.style) o.style.display = 'none'; });
                        b.click(); return true;
                    }
                }
                return false;
            }"""
        )
        if still_open:
            await asyncio.sleep(2)
        else:
            break
    await asyncio.sleep(3)

    # 切维度
    dim_idx = DIMENSION_MAP.get(dimension, 1)
    dim_label = {v: k for k, v in DIMENSION_MAP.items()}.get(dim_idx, dimension)
    if dim_idx != 1:
        print(f"  切维度: {dim_label} (index {dim_idx})...")
        # 等维度容器出现
        for _ in range(30):
            ready = await page.evaluate(
                """() => {
                    const all = document.querySelectorAll('*');
                    for (const el of all) {
                        if (el.children.length > 3 && el.offsetWidth > 0) {
                            const t = el.innerText;
                            if (t.includes('自然排名') && t.includes('广告排名') && t.length < 300) {
                                return true;
                            }
                        }
                    }
                    return false;
                }"""
            )
            if ready:
                break
            await asyncio.sleep(1)
        await page.evaluate(
            """(idx) => {
                const all = document.querySelectorAll('*');
                for (const el of all) {
                    if (el.children.length > 3 && el.offsetWidth > 0) {
                        const t = el.innerText;
                        if (t.includes('自然排名') && t.includes('广告排名') && t.length < 300) {
                            el.children[idx].click(); return true;
                        }
                    }
                }
                return false;
            }""",
            dim_idx,
        )
        await asyncio.sleep(5)
        print("  维度切换完成")
    else:
        print(f"  维度: {dim_label} (默认)")

    # 等结果加载：检查导出按钮是否出现并可用
    print("  等待数据加载...", end="", flush=True)
    for _ in range(120):
        ready = await page.evaluate(
            """() => {
                const btns = document.querySelectorAll('button');
                let hasExport = false;
                for (const b of btns) {
                    if (b.offsetWidth > 0) {
                        const t = b.innerText.trim();
                        if (t === '导出' && !b.disabled) {
                            hasExport = true;
                        }
                    }
                }
                if (hasExport && document.querySelectorAll('table tbody tr').length > 3) {
                    return true;
                }
                return false;
            }"""
        )
        if ready:
            print(" 完成")
            return True
        await asyncio.sleep(2)
        print(".", end="", flush=True)
    print(" 超时")
    return False


async def trigger_export(page) -> bool:
    """点击导出 + 前往查看"""
    print("  导出...", end="", flush=True)
    clicked = await page.evaluate(
        """() => {
            const btns = document.querySelectorAll('button');
            for (const b of btns) {
                if (b.offsetWidth > 0 && !b.disabled && b.innerText.trim() === '导出') {
                    b.click(); return true;
                }
            }
            return false;
        }"""
    )
    if not clicked:
        print(" 失败")
        return False
    await asyncio.sleep(2)

    # 等「前往查看」
    await asyncio.sleep(1)
    await page.evaluate(
        """() => {
            const btns = document.querySelectorAll('button');
            for (const b of btns) {
                if (b.offsetWidth > 0 && b.innerText.includes('前往查看')) {
                    b.click(); return;
                }
            }
        }"""
    )
    await asyncio.sleep(2)
    print(" 已提交")
    return True


async def poll_and_download(page, output_dir: Path, label: str) -> Path | None:
    """轮询导出记录 → 下载"""
    # 找导出记录页
    export_page = None
    for _ in range(10):
        for p in page.context.pages:
            if "export-log" in p.url:
                export_page = p
                break
        if export_page:
            break
        await asyncio.sleep(1)
    if not export_page:
        export_page = page
        await export_page.goto(EXPORT_LOG_URL, timeout=30000)
        await export_page.wait_for_load_state("domcontentloaded")
        await asyncio.sleep(2)

    # 找任务名
    task_name = await export_page.evaluate(
        """() => {
            const rows = document.querySelectorAll('tr');
            for (const r of rows) {
                if (r.innerText.includes('CompareKeywords-') && r.innerText.includes('流量词对比')) {
                    const m = r.innerText.match(/CompareKeywords-[A-Z]+-[A-Za-z0-9]+-\\d+-\\d+/);
                    if (m) return m[0];
                }
            }
            return null;
        }"""
    )
    if not task_name:
        print("  找不到导出任务")
        return None
    print(f"  任务: {task_name}")

    # 轮询
    print("  等待完成", end="", flush=True)
    for _ in range(60):
        await asyncio.sleep(10)
        await export_page.reload(timeout=30000)
        await export_page.wait_for_load_state("domcontentloaded")
        await asyncio.sleep(0.5)
        status = await export_page.evaluate(
            """(n) => {
                const rows = document.querySelectorAll('tr');
                for (const r of rows) {
                    if (r.innerText.includes(n)) {
                        if (r.innerText.includes('已完成')) return 'done';
                        if (r.innerText.includes('导出中')) return 'exporting';
                    }
                }
                return '?';
            }""",
            task_name,
        )
        if status == "done":
            print(" 完成")
            break
        if status == "?":
            print(" 丢失")
            return None
        print(".", end="", flush=True)

    # 下载
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{task_name}.xlsx"
    try:
        df = asyncio.get_event_loop().create_future()

        async def on_download(dl):
            await dl.save_as(str(output_path))
            df.set_result(output_path)

        export_page.on("download", on_download)
        await export_page.evaluate(
            """(n) => {
                const rows = document.querySelectorAll('tr');
                for (const r of rows) {
                    if (r.innerText.includes(n)) {
                        const links = r.querySelectorAll('a[href*=".xlsx"]');
                        for (const l of links) { if (l.offsetWidth > 0) { l.click(); return; } }
                    }
                }
            }""",
            task_name,
        )
        path = await asyncio.wait_for(df, timeout=120)
        print(f"  已保存: {path}")
        return path
    except asyncio.TimeoutError:
        print("  下载超时")
        return None
    except Exception as e:
        print(f"  下载失败: {e}")
        return None
    finally:
        try:
            export_page.remove_listener("download", on_download)
        except Exception:
            pass


# ============================================================
#  清洗
# ============================================================

def rank_tier(label: str) -> str:
    if not label:
        return "-"
    s = str(label).strip()
    if s == PAGE1_RANK_LABEL:
        return "P1"
    if s == "精准流量词":
        return "P2"
    if s == "精准长尾词":
        return "P3"
    return s


def parse_traffic_matrix(xlsx_path: str | Path) -> dict:
    """解析交叉矩阵"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    asins = []
    for j in range(3, ws.max_column + 1, 2):
        v = ws.cell(1, j).value
        if v and str(v).strip() and len(str(v).strip()) >= 10:
            asins.append(str(v).strip())

    agg_start = 3 + len(asins) * 2
    rows = []
    for row in range(2, ws.max_row + 1):
        kw = ws.cell(row, 1).value
        if not kw:
            continue
        kw = str(kw).strip()
        sv = ws.cell(row, agg_start + 2).value or ""
        ad = {}
        for i, asin in enumerate(asins):
            sc = 3 + i * 2
            ad[asin] = {
                "share": ws.cell(row, sc).value or "",
                "tier": rank_tier(ws.cell(row, sc + 1).value),
            }
        rows.append({"keyword": kw, "search_volume": sv, "asins": ad})
    return {"asins": asins, "rows": rows}


def parse_natural_rank_matrix(xlsx_path: str | Path) -> dict:
    """解析自然排名导出表"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 每个 ASIN 3 列：排名位置 / 页码 / 抓取时间
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        headers[col] = str(v).strip() if v else ""

    asin_cols = []  # [(asin, rank_col, page_col, time_col)]
    last_asin_end = 2
    for col, h in headers.items():
        if col <= last_asin_end:
            continue
        if len(h) >= 10 and h[:2] == "B0" and col > 2:
            asin_cols.append((h, col, col + 1, col + 2))
            last_asin_end = col + 2

    if not asin_cols:
        # 兼容旧格式（每 ASIN 2 列）
        return parse_traffic_matrix(xlsx_path)

    # 聚合列：关键词(1) + 翻译(2) + N*3(ASIN) 之后
    agg_start = 2 + len(asin_cols) * 3

    rows = []
    for row in range(2, ws.max_row + 1):
        kw = ws.cell(row, 1).value
        if not kw:
            continue
        kw = str(kw).strip()
        sv = ws.cell(row, agg_start + 2).value or ""  # 搜索量
        ad = {}
        for asin, rc, pc, tc in asin_cols:
            rank_val = ws.cell(row, rc).value
            page_val = ws.cell(row, pc).value
            ad[asin] = {
                "rank": str(rank_val).strip() if rank_val else "-",
                "page": str(page_val).strip() if page_val else "-",
            }
        rows.append({"keyword": kw, "search_volume": sv, "asins": ad})
    return {"asins": [a for a, _, _, _ in asin_cols], "rows": rows}


def is_page1(page_str: str) -> bool:
    """判断是否第1页"""
    return "第1页" in page_str


def save_p1_organic_report(matrix: dict, output_path: Path, top_n: int = 20):
    """输出每个 ASIN 第1页自然排名关键词"""
    asins = matrix["asins"]
    p1_data: dict[str, list[dict]] = defaultdict(list)
    for rd in matrix["rows"]:
        for asin in asins:
            a = rd["asins"][asin]
            if is_page1(a["page"]):
                p1_data[asin].append({
                    "keyword": rd["keyword"],
                    "rank": a["rank"],
                    "page_info": a["page"],
                    "search_volume": rd["search_volume"],
                })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P1自然排名"
    for j, h in enumerate(["ASIN", "关键词", "自然排名", "页码", "月搜索量"], 1):
        ws.cell(1, j, h)
    r = 2
    for asin in asins:
        entries = p1_data.get(asin, [])
        entries.sort(key=lambda e: float(e["rank"]) if e["rank"].replace(".", "").isdigit() else 9999)
        for e in entries:
            ws.cell(r, 1, asin)
            ws.cell(r, 2, e["keyword"])
            ws.cell(r, 3, e["rank"])
            ws.cell(r, 4, e["page_info"])
            ws.cell(r, 5, e["search_volume"])
            r += 1
    wb.save(output_path)

    for asin in asins:
        entries = p1_data.get(asin, [])
        print(f"  {asin}: {len(entries)} 第1页关键词")
        for e in entries[:top_n]:
            print(f"    #{e['rank']} {e['keyword']} (搜索量:{e['search_volume']})")


TIER_LABELS = {"P1": "主要流量词", "P2": "精准流量词", "P3": "精准长尾词"}

def save_tier_report(matrix: dict, output_path: Path, tier: str):
    """输出指定 tier 的流量入口"""
    asins = matrix["asins"]
    tname = TIER_LABELS.get(tier, tier)
    data: dict[str, list[dict]] = defaultdict(list)
    for rd in matrix["rows"]:
        for asin in asins:
            a = rd["asins"][asin]
            if tier == "all" or a["tier"] == tier:
                data[asin].append({
                    "keyword": rd["keyword"],
                    "share": a["share"],
                    "search_volume": rd["search_volume"],
                    "tier": a["tier"],
                })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{tier}流量入口" if tier != "all" else "全层级流量入口"
    headers = ["ASIN", "关键词", "流量占比", "月搜索量"]
    if tier == "all":
        headers.insert(3, "层级")
    for j, h in enumerate(headers, 1):
        ws.cell(1, j, h)
    r = 2
    for asin in asins:
        for e in data.get(asin, []):
            ws.cell(r, 1, asin)
            ws.cell(r, 2, e["keyword"])
            ws.cell(r, 3, e["share"])
            if tier == "all":
                ws.cell(r, 4, e["tier"])
                ws.cell(r, 5, e["search_volume"])
            else:
                ws.cell(r, 4, e["search_volume"])
            r += 1
    wb.save(output_path)

    for asin in asins:
        entries = data.get(asin, [])
        print(f"  {asin}: {len(entries)} {tier}")
        for e in entries[:20]:
            extra = f" [{e['tier']}]" if tier == "all" else ""
            print(f"    {e['keyword']} ({e['share']}){extra}")


# ============================================================
#  主流程
# ============================================================

async def main():
    parser = argparse.ArgumentParser(description="卖家精灵流量词对比 → ASIN P1 流量入口")
    parser.add_argument("asins_file", help="ASIN 文件，每行一个（最多 10 个）")
    parser.add_argument("--market", default="US")
    parser.add_argument("--own-asin", default="")
    parser.add_argument("--dimension", default="traffic_share", choices=list(DIMENSION_MAP.keys()))
    parser.add_argument("--prefix", default="")
    parser.add_argument("--output", default=".")
    parser.add_argument("--top", type=int, default=20, help="打印前N个关键词")
    parser.add_argument("--tier", default="P1", choices=["P1", "P2", "P3", "all"],
                        help="输出层级: P1主要流量词 / P2精准流量词 / P3精准长尾词 / all全部")
    args = parser.parse_args()

    ap = Path(args.asins_file)
    if not ap.exists():
        print(f"文件不存在: {ap}")
        sys.exit(1)

    asins = list(dict.fromkeys(
        [l.strip() for l in ap.read_text(encoding="utf-8").splitlines() if l.strip()]
    ))
    if not asins:
        print("ASIN 文件为空")
        sys.exit(1)
    if len(asins) > MAX_ASINS:
        print(f"超过 {MAX_ASINS} 个，取前 {MAX_ASINS} 个")
        asins = asins[:MAX_ASINS]

    print(f"{len(asins)} 个 ASIN, 市场: {args.market}, 维度: {args.dimension}")
    output_dir = Path(args.output).resolve()
    now_str = time.strftime("%Y%m%d-%H%M%S")
    prefix = args.prefix or f"CMP-{asins[0]}"
    label = f"{prefix}-{args.market}-{now_str}"

    USER_DATA.mkdir(parents=True, exist_ok=True)

    async with async_playwright() as p:
        try:
            ctx = await p.chromium.launch_persistent_context(
                user_data_dir=str(USER_DATA),
                headless=False,
                args=["--no-sandbox", "--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            import shutil
            shutil.rmtree(str(USER_DATA), ignore_errors=True)
            USER_DATA.mkdir(parents=True, exist_ok=True)
            ctx = await p.chromium.launch_persistent_context(
                user_data_dir=str(USER_DATA),
                headless=False,
                args=["--no-sandbox", "--disable-blink-features=AutomationControlled"],
            )

        page = await ctx.new_page()
        await ensure_logged_in(page)

        if not await run_comparison(page, asins, args.own_asin, args.dimension):
            print("查询失败")
            sys.exit(1)
        if not await trigger_export(page):
            print("导出失败")
            sys.exit(1)

        raw = await poll_and_download(page, output_dir, label)
        if not raw:
            print("下载失败")
            sys.exit(1)

        await ctx.close()

    # 清洗
    if args.dimension == "natural_rank":
        print("\n解析自然排名 → 第1页关键词...")
        matrix = parse_natural_rank_matrix(raw)
        report = output_dir / f"P1-Organic-{label}.xlsx"
        save_p1_organic_report(matrix, report, top_n=args.top)
    else:
        tname = TIER_LABELS.get(args.tier, args.tier)
        print(f"\n解析 {tname}({args.tier}) 流量入口...")
        matrix = parse_traffic_matrix(raw)
        report = output_dir / f"{args.tier}-{label}.xlsx"
        save_tier_report(matrix, report, args.tier)
    print(f"\n原始: {raw}")
    print(f"报告: {report}")


if __name__ == "__main__":
    asyncio.run(main())
