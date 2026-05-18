"""
需求匹配指数分析工具
用法：
  1. 双击 exe 运行，在控制台输入 xlsx 文件路径
  2. 或命令行: demand_match_index.exe <文件路径>

计算公式：需求匹配指数 = 转化份额 / 点击份额
输出：需求匹配指数 > 1 的 ASIN 及其关联关键词信息
"""

import sys
import os
import openpyxl
from datetime import datetime


def analyze(filepath):
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return

    print(f"\n正在分析: {filepath}")
    print("=" * 80)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]

    # 读取数据行
    results = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not row or not row[1]:
            continue

        ranking = row[0]       # 搜索词排名
        search_term = row[1]   # 搜索词
        report_date = row[20]  # 报告日期

        # 品牌/类目信息
        brand1 = row[2]
        brand2 = row[3]
        brand3 = row[4]
        cat1 = row[5]
        cat2 = row[6]
        cat3 = row[7]

        # 处理 3 个 ASIN
        for slot in range(3):
            base = 8 + slot * 4
            asin = row[base]
            product_name = row[base + 1]
            click_share = row[base + 2]
            conversion_share = row[base + 3]
            brand = [brand1, brand2, brand3][slot]
            category = [cat1, cat2, cat3][slot]

            if not asin:
                continue

            # 处理空值和零值
            click_share = float(click_share) if click_share else 0
            conversion_share = float(conversion_share) if conversion_share else 0

            # 点击份额为 0 则跳过（无法计算指数）
            if click_share == 0:
                continue

            demand_index = conversion_share / click_share

            if demand_index > 1:
                results.append({
                    "ranking": ranking,
                    "search_term": search_term,
                    "report_date": report_date,
                    "asin": asin,
                    "product_name": product_name,
                    "brand": brand,
                    "category": category,
                    "click_share": click_share,
                    "conversion_share": conversion_share,
                    "demand_index": round(demand_index, 4),
                    "slot": slot + 1,
                })

    # 按需求匹配指数降序排序
    results.sort(key=lambda x: x["demand_index"], reverse=True)

    if not results:
        print("\n没有找到需求匹配指数 > 1 的 ASIN")
        return

    # 控制台输出
    print(f"\n找到 {len(results)} 个需求匹配指数 > 1 的 ASIN:\n")
    print(f"{'排名':<6}{'搜索词':<30}{'ASIN':<14}{'商品名称':<30}{'点击份额':<10}{'转化份额':<10}{'需求匹配指数':<14}{'品牌':<16}{'类目':<20}")
    print("-" * 150)

    for r in results:
        name = str(r["product_name"])[:28] if r["product_name"] else ""
        term = str(r["search_term"])[:28] if r["search_term"] else ""
        brand = str(r["brand"])[:14] if r["brand"] else ""
        cat = str(r["category"])[:18] if r["category"] else ""
        print(f"{r['ranking']:<6}{term:<30}{r['asin']:<14}{name:<30}{r['click_share']:<10}{r['conversion_share']:<10}{r['demand_index']:<14}{brand:<16}{cat:<20}")

    # 输出 Excel
    output_path = os.path.splitext(filepath)[0] + "_需求匹配指数结果.xlsx"
    owb = openpyxl.Workbook()
    ows = owb.active
    ows.title = "需求匹配指数>1"

    out_headers = [
        "搜索词排名", "搜索词", "报告日期",
        "ASIN", "商品名称", "品牌", "类目",
        "点击份额", "转化份额", "需求匹配指数", "排名位次(#1/#2/#3)"
    ]
    ows.append(out_headers)

    for r in results:
        date_str = r["report_date"].strftime("%Y-%m-%d") if isinstance(r["report_date"], datetime) else str(r["report_date"])
        ows.append([
            r["ranking"], r["search_term"], date_str,
            r["asin"], r["product_name"], r["brand"], r["category"],
            r["click_share"], r["conversion_share"], r["demand_index"],
            f"#{r['slot']}"
        ])

    # 设置列宽
    col_widths = [12, 30, 14, 16, 50, 18, 22, 12, 12, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ows.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    owb.save(output_path)
    print(f"\n结果已保存到: {output_path}")
    print("=" * 80)


def main():
    print("需求匹配指数分析工具")
    print("公式: 需求匹配指数 = 转化份额 / 点击份额")
    print("筛选: 需求匹配指数 > 1 的 ASIN\n")

    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        filepath = input("请输入 xlsx 文件路径（可拖拽文件到窗口）: ").strip().strip('"').strip("'")

    analyze(filepath)
    try:
        input("\n按回车键退出...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
